"""Flask API for the media personalisation pipeline.

POST /upload-template  multipart {file, kind: 'video'|'image'} -> saves template,
                       updates config/settings.json so future generators use it.
POST /generate         { address, phone }                     -> renders ONE pair
                       (image + video) for a single recipient.
POST /generate-videos  { recipients: [{name?, address, phone}, ...] }
                       -> loops the list, renders image+video per row, returns
                       a per-row result list.
GET  /health                                                   -> { status: "ok" }

Designed to run from any cwd; resolves scripts and project files from
__file__ so `python api/app.py`, `python -m api.app`, and `cd api && python app.py`
all behave identically.
"""

import csv
import hashlib
import hmac
import io
import json
import os
import re
import subprocess
import sys
import threading
import time
import uuid
from datetime import datetime

# --------------------------------------------------------------------------
# Early .env load — runs BEFORE any module-level os.environ.get() so the
# provider auto-bootstrap, CORS origins, SESSION_SECRET, and ProxyFix all
# see the values from api/.env. The full-featured loader (with logging)
# lives at the bottom of this file for legacy compatibility; this tiny
# stub does the same job earlier.
#
# In Vercel serverless: skip file loading, environment variables come from
# the Vercel dashboard instead. VERCEL=1 is set by the platform.
# --------------------------------------------------------------------------
def _early_load_env():
    # Skip file loading in Vercel serverless — use dashboard env vars instead
    if os.environ.get("VERCEL") == "1":
        print("[app] Vercel serverless detected, skipping .env file load (using dashboard env vars)", flush=True)
        return
    
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env.vercel")
    if not os.path.isfile(env_path):
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.isfile(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith("#") or "=" not in s:
                    continue
                k, v = s.split("=", 1)
                k, v = k.strip(), v.strip()
                if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
                    v = v[1:-1]
                if k and k not in os.environ:
                    os.environ[k] = v
    except Exception:  # pragma: no cover  — fall through silently in dev
        pass

_early_load_env()

# Optional dependencies — degrade gracefully if either is missing.
try:
    import requests as _requests
except ImportError:
    _requests = None
try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from werkzeug.middleware.proxy_fix import ProxyFix

# Local — provider adapter layer. Keep this file in the same dir as app.py.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from providers import get_provider, list_providers  # noqa: E402
import session as _session  # noqa: E402  shared HMAC session module

PROJECT_ROOT     = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
SCRIPT_IMAGE     = os.path.join(PROJECT_ROOT, "scripts", "image_generator.py")
SCRIPT_VIDEO    = os.path.join(PROJECT_ROOT, "scripts", "video_generator.py")
if os.environ.get("VERCEL") == "1":
    DATA_ROOT = "/tmp"
else:
    DATA_ROOT = PROJECT_ROOT
TEMPLATES_DIR    = os.path.join(DATA_ROOT, "templates")
OUTPUT_VIDEOS    = os.path.join(DATA_ROOT, "output", "videos")
OUTPUT_IMAGES    = os.path.join(DATA_ROOT, "output", "images")
SETTINGS_PATH    = os.path.join(DATA_ROOT, "config", "settings.json")
RECIPIENTS_PATH  = os.path.join(DATA_ROOT, "data", "recipients.json")
SHEET_SOURCE_PATH = os.path.join(DATA_ROOT, "data", "sheet-source.json")
DELIVERIES_PATH  = os.path.join(DATA_ROOT, "data", "deliveries.json")
DELIVERY_LOG_PATH = os.path.join(DATA_ROOT, "data", "delivery-logs.jsonl")
JOBS_DIR         = os.path.join(DATA_ROOT, "data", "jobs")
if os.environ.get("VERCEL") == "1":
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    os.makedirs(OUTPUT_VIDEOS, exist_ok=True)
    os.makedirs(OUTPUT_IMAGES, exist_ok=True)
    os.makedirs(os.path.dirname(SETTINGS_PATH), exist_ok=True)
    os.makedirs(os.path.dirname(RECIPIENTS_PATH), exist_ok=True)
    os.makedirs(JOBS_DIR, exist_ok=True)
    if not os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH, "w") as f: f.write("{}")
else:
    os.makedirs(JOBS_DIR, exist_ok=True)
TIMEOUT_SECONDS  = 180  # full image+video render ceiling
MAX_UPLOAD_BYTES = 200 * 1024 * 1024  # matches the frontend dropzone copy

# Extensions we recognise as "generated outputs" — anything else in the folder
# (e.g. leftover .DS_Store, partial files) is ignored.
_VIDEO_OUT_EXTS = {".mp4", ".mov", ".webm", ".m4v"}
_IMAGE_OUT_EXTS = {".png", ".jpg", ".jpeg", ".webp"}

# Match the line each generator prints on success.
_PATH_RE = re.compile(r"^Generated (?:image|video):\s*(.+)$", re.MULTILINE)

app = Flask(__name__)
# Flask's own request-body cap — bigger than the upload to leave headroom for
# multipart boundaries.
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES + (4 * 1024 * 1024)

# Reverse-proxy header support — when this Flask process sits behind nginx
# / Caddy / Cloudflare, request.remote_addr should be the real client IP
# (X-Forwarded-For), request.scheme should reflect the public scheme
# (X-Forwarded-Proto), and request.host should be the public host
# (X-Forwarded-Host). Trust exactly one proxy hop by default; bump via
# TRUSTED_PROXIES env if you have multiple.
try:
    _trusted_proxies = int(os.environ.get("TRUSTED_PROXIES", "1"))
except (TypeError, ValueError):
    _trusted_proxies = 1
app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=_trusted_proxies,
    x_proto=_trusted_proxies,
    x_host=_trusted_proxies,
    x_port=_trusted_proxies,
)

# CORS origin whitelist. Defaults to the Vite dev server only; production
# deploys MUST set CORS_ORIGINS to the real frontend host (comma-separated
# for multiple). Wildcard `*` is intentionally NOT the default — leaving
# this open is what the pre-deployment audit flagged as a critical risk.
_cors_origins_env = os.environ.get("CORS_ORIGINS", "http://localhost:5173").strip()
_cors_origins = [o.strip() for o in _cors_origins_env.split(",") if o.strip()]
CORS(app, resources={r"/*": {"origins": _cors_origins}})
print(f"[startup] CORS origins: {_cors_origins}", flush=True)

if os.environ.get("VERCEL") == "1":
    class VercelPrefixMiddleware:
        def __init__(self, app, prefix):
            self.app = app
            self.prefix = prefix
        def __call__(self, environ, start_response):
            path = environ.get("PATH_INFO", "")
            if path.startswith(self.prefix):
                environ["PATH_INFO"] = path[len(self.prefix):]
                environ["SCRIPT_NAME"] = self.prefix
            return self.app(environ, start_response)
    app.wsgi_app = VercelPrefixMiddleware(app.wsgi_app, "/api")



# =============================================================================
# Authentication middleware
#
# Every request to this Flask process is gated unless it matches one of the
# public prefixes. Public set is intentionally tiny:
#
#   /health                       - liveness probe (no data)
#   /files/<videos|images>/...    - generated media served by URL; the
#                                   <img>/<video> tags can't set Auth
#                                   headers, so URL guessability is the
#                                   only protection. Filenames are derived
#                                   from sanitised addresses, so they're
#                                   not user-enumerable but ARE leakable
#                                   if a URL escapes (e.g. via a copied
#                                   share link). DOCUMENTED RISK.
#   OPTIONS (CORS preflight)     - browsers send these without credentials.
#
# Everything else requires `Authorization: Bearer <token>` where <token>
# is a valid HMAC-signed session token issued by auth_server.py. Both
# servers share `SESSION_SECRET` so verification needs no DB lookup.
# =============================================================================
_PUBLIC_PREFIXES = (
    "/health",
    "/files/videos/",
    "/files/images/",
    # WhatsApp's webhook calls from Meta's network — can't carry a Bearer
    # token. Cryptographically verified instead via X-Hub-Signature-256
    # inside the handler (fail-closed: rejects with 403 if WHATSAPP_APP_SECRET
    # isn't set).
    "/deliveries/whatsapp-webhook",
)


@app.before_request
def _require_auth():
    # CORS preflight requests must always pass through unauthenticated.
    if request.method == "OPTIONS":
        return None
    # Public allowlist (prefix match).
    path = request.path or ""
    if any(path == p or path.startswith(p) for p in _PUBLIC_PREFIXES):
        return None
    # Everything else: verify Bearer token.
    token = _session.extract_bearer(request.headers.get("Authorization"))
    if not token:
        return jsonify({"status": "error", "error": "auth required"}), 401
    payload = _session.verify_token(token)
    if not payload:
        return jsonify({"status": "error", "error": "invalid or expired token"}), 401
    # Stash the verified user on the Flask request context for handlers
    # that want to log who did what.
    request.user_id = payload["user_id"]
    return None


# Global error handler for unhandled exceptions
@app.errorhandler(Exception)
def handle_error(e):
    """Catch all unhandled exceptions and return JSON error response."""
    import traceback
    print(f"[ERROR] Unhandled exception: {str(e)}\n{traceback.format_exc()}", flush=True)
    return jsonify({
        "status": "error",
        "error": "internal_error",
        "details": str(e)
    }), 500


def _run_generator(script_path, address, phone, name=""):
    """Run a single generator script and return a structured result dict.

    The generators accept argv[1]=address, argv[2]=phone, argv[3]=name
    (optional). When name is empty the generator falls back to the
    legacy two-line layout, preserving backward compatibility."""
    cmd = [sys.executable, script_path, address, phone]
    if name:
        cmd.append(name)
    try:
        proc = subprocess.run(
            cmd,
            cwd=PROJECT_ROOT,
            capture_output=True,
            text=True,
            timeout=TIMEOUT_SECONDS,
        )
    except subprocess.TimeoutExpired as e:
        return {
            "ok": False,
            "stdout": e.stdout or "",
            "stderr": f"timed out after {TIMEOUT_SECONDS}s",
            "path": None,
            "returncode": None,
        }
    except FileNotFoundError as e:
        return {
            "ok": False,
            "stdout": "",
            "stderr": f"script not found: {e}",
            "path": None,
            "returncode": None,
        }

    stdout = proc.stdout or ""
    stderr = proc.stderr or ""
    match = _PATH_RE.search(stdout)
    return {
        "ok": proc.returncode == 0,
        "stdout": stdout,
        "stderr": stderr,
        "path": match.group(1).strip() if match else None,
        "returncode": proc.returncode,
    }


# -------------------------------------------------------------------------
# Pipeline-specific entry points.
#
# These are the ONLY two functions in the app that should ever shell out to
# the renderer scripts. Each one is hard-bound to a single script and a
# single output directory. There is no shared "render both" function; if
# you want both formats you call both functions explicitly (the legacy
# kind='all' path is the only place that does so, and it's clearly marked).
# -------------------------------------------------------------------------
def _run_image_pipeline(address, phone, name=""):
    """IMAGE PIPELINE -- runs ONLY the image generator. Never invokes
    ffmpeg or the video generator. Produces a PNG in output/images/."""
    print(f"  -> _run_image_pipeline(addr={address!r}) -- invoking SCRIPT_IMAGE only", flush=True)
    return _run_generator(SCRIPT_IMAGE, address, phone, name)


def _run_video_pipeline(address, phone, name=""):
    """VIDEO PIPELINE -- runs ONLY the video generator (ffmpeg + libass).
    Never invokes the image generator. Produces an MP4 in output/videos/."""
    print(f"  -> _run_video_pipeline(addr={address!r}) -- invoking SCRIPT_VIDEO only", flush=True)
    return _run_generator(SCRIPT_VIDEO, address, phone, name)


def _load_settings():
    with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_settings(settings):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2)
        f.write("\n")


@app.route("/health", methods=["GET"])
def health():
    """Liveness + lightweight observability. No auth required — this is
    what monitoring / the tunnel keep-alive / Cloudflare healthchecks hit.
    Body is intentionally minimal (no secrets, just status and config flags
    that help an operator confirm the right process is responding)."""
    return jsonify({
        "status":               "ok",
        "provider":             _PROVIDER_NAME,
        "time":                 datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "public_base_url":      (os.environ.get("PUBLIC_BASE_URL") or "").strip() or None,
        "verify_token_present": bool((os.environ.get("WHATSAPP_WEBHOOK_VERIFY_TOKEN") or "").strip()),
        "app_secret_present":   bool((os.environ.get("WHATSAPP_APP_SECRET")          or "").strip()),
    })


@app.route("/generate", methods=["POST"])
def generate():
    """Synchronous single-recipient render. Honors a `kind` parameter so
    image and video pipelines stay strictly independent — the Dashboard
    quick-action modal lets users pick exactly one kind. Backwards
    compatible: omitting `kind` defaults to "image" only (the safer choice
    than the old "always render both" behavior, which silently doubled
    work)."""
    payload = request.get_json(silent=True) or {}
    address = str(payload.get("address") or "").strip()
    phone   = str(payload.get("phone") or "").strip()
    name    = str(payload.get("name") or "").strip()
    kind    = (payload.get("kind") or "image").strip().lower()

    if not address or not phone:
        return jsonify({
            "status": "error",
            "error": "address and phone are required",
        }), 400
    if kind not in ("image", "video"):
        return jsonify({
            "status": "error",
            "error": "kind must be 'image' or 'video'",
        }), 400

    print(f"[/generate] MEDIA TYPE={kind.upper()} name={name!r} address={address!r} phone={phone!r}", flush=True)

    image = None
    video = None
    if kind == "image":
        image = _run_image_pipeline(address, phone, name)
    else:
        video = _run_video_pipeline(address, phone, name)

    overall_ok = (image["ok"] if image else True) and (video["ok"] if video else True)
    body = {
        "status": "success" if overall_ok else "error",
        "kind":   kind,
    }
    if image is not None:
        body["image"] = image
    if video is not None:
        body["video"] = video
    if not overall_ok:
        stderr = (image and image["stderr"]) or (video and video["stderr"]) or ""
        body["error"] = stderr.strip().splitlines()[-1] if stderr.strip() else "generator failed"
    return jsonify(body), (200 if overall_ok else 500)


# ---------------------------------------------------------------------------
# Template upload
# ---------------------------------------------------------------------------

# Whitelisted file extensions per kind. We do NOT trust the client MIME alone —
# we accept based on extension *and* MIME prefix.
_VIDEO_EXTS = {".mp4", ".mov", ".webm", ".m4v"}
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}


@app.route("/upload-template", methods=["POST"])
def upload_template():
    kind = (request.form.get("kind") or "").strip().lower()
    if kind not in ("video", "image"):
        return jsonify({"status": "error", "error": "kind must be 'video' or 'image'"}), 400

    if "file" not in request.files:
        return jsonify({"status": "error", "error": "no file part in request"}), 400
    f = request.files["file"]
    if not f or not (f.filename or "").strip():
        return jsonify({"status": "error", "error": "empty filename"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    allowed = _VIDEO_EXTS if kind == "video" else _IMAGE_EXTS
    if ext not in allowed:
        return jsonify({
            "status": "error",
            "error": f"unsupported {kind} extension {ext!r}; allowed: {sorted(allowed)}",
        }), 400

    mt = (f.mimetype or "").lower()
    expected_prefix = "video/" if kind == "video" else "image/"
    if mt and not mt.startswith(expected_prefix):
        return jsonify({
            "status": "error",
            "error": f"expected {expected_prefix}*, got {mt!r}",
        }), 400

    # Probe size cheaply without loading the whole stream into memory.
    f.stream.seek(0, os.SEEK_END)
    size = f.stream.tell()
    f.stream.seek(0)
    if size <= 0:
        return jsonify({"status": "error", "error": "file is empty"}), 400
    if size > MAX_UPLOAD_BYTES:
        return jsonify({
            "status": "error",
            "error": f"file is {size} bytes; max {MAX_UPLOAD_BYTES} bytes (200 MB)",
        }), 413

    # Read file into memory and encode as base64 for persistent storage in settings.json
    file_data = f.read()
    file_b64 = __import__('base64').b64encode(file_data).decode('ascii')
    
    # Also save to temp location for immediate availability in generators
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    target_name = f"uploaded_{kind}{ext}"
    target_path = os.path.join(TEMPLATES_DIR, target_name)
    with open(target_path, 'wb') as fp:
        fp.write(file_data)

    saved_size = os.path.getsize(target_path)
    rel_path = f"templates/{target_name}".replace("\\", "/")

    # Update settings.json with both the path AND the base64-encoded file content
    # so it survives Vercel serverless invocations
    settings = _load_settings()
    template_key = f"template_{kind}"
    data_key = f"template_{kind}_data"
    filename_key = f"template_{kind}_filename"
    
    settings[template_key] = rel_path
    settings[data_key] = file_b64  # base64-encoded file content
    settings[filename_key] = f.filename  # original filename
    
    _save_settings(settings)

    print(
        f"[/upload-template] kind={kind} saved={target_path} bytes={saved_size} "
        f"-> settings.template_{kind} = {rel_path} (with base64 fallback)",
        flush=True,
    )

    return jsonify({
        "status": "success",
        "kind": kind,
        "filename": f.filename,
        "saved_as": target_name,
        "path": rel_path,
        "abs_path": target_path,
        "bytes": saved_size,
    })


# ---------------------------------------------------------------------------
# Recipients (Google-Sheets-like store) — persisted to data/recipients.json
# ---------------------------------------------------------------------------

# Lock so concurrent writers can't corrupt the JSON.
_RECIPIENTS_LOCK = threading.Lock()

# Validation regexes. Phone: optional +, then 7-20 chars that look like
# digits/spaces/dashes/parens. Final check counts pure digits.
_PHONE_ALLOWED = re.compile(r"^\+?[\d\s\-\(\)]+$")
_PHONE_MIN_DIGITS = 7
_PHONE_MAX_DIGITS = 15  # ITU-T E.164 max


def _normalize_phone(phone):
    """Return digit-only form of the phone for duplicate detection."""
    return re.sub(r"\D", "", phone or "")


def _normalize_address(addr):
    return re.sub(r"\s+", " ", (addr or "").strip()).lower()


def _empty_recipients_doc():
    return {"version": 1, "updatedAt": 0, "items": []}


def _load_recipients():
    """Read the recipients JSON. Falls back to an empty doc when missing /
    corrupt — never raises into the request handler."""
    if not os.path.isfile(RECIPIENTS_PATH):
        return _empty_recipients_doc()
    try:
        with open(RECIPIENTS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        return _empty_recipients_doc()
    if not isinstance(data, dict):
        return _empty_recipients_doc()
    items = data.get("items") or []
    if not isinstance(items, list):
        items = []
    # Keep only well-shaped entries.
    clean_items = []
    for it in items:
        if not isinstance(it, dict):
            continue
        clean_items.append({
            "id":      str(it.get("id") or _new_id()),
            "name":    str(it.get("name") or "").strip(),
            "phone":   str(it.get("phone") or "").strip(),
            "address": str(it.get("address") or "").strip(),
        })
    return {
        "version":   int(data.get("version") or 1),
        "updatedAt": int(data.get("updatedAt") or 0),
        "items":     clean_items,
    }


def _save_recipients(doc):
    doc["updatedAt"] = int(time.time() * 1000)
    os.makedirs(os.path.dirname(RECIPIENTS_PATH), exist_ok=True)
    tmp = RECIPIENTS_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(doc, f, indent=2, ensure_ascii=False)
        f.write("\n")
    os.replace(tmp, RECIPIENTS_PATH)


def _new_id():
    return f"r_{uuid.uuid4().hex[:10]}"


def _validate_row(row, seen_phones, seen_addresses):
    """Return list of error strings for this row. Mutates seen_* sets."""
    errors = []
    name    = (row.get("name") or "").strip()
    phone   = (row.get("phone") or "").strip()
    address = (row.get("address") or "").strip()

    if not name:
        errors.append("missing name")
    if not phone:
        errors.append("missing phone")
    else:
        digits = _normalize_phone(phone)
        if not _PHONE_ALLOWED.match(phone):
            errors.append("invalid phone characters")
        elif len(digits) < _PHONE_MIN_DIGITS:
            errors.append(f"phone has only {len(digits)} digits (min {_PHONE_MIN_DIGITS})")
        elif len(digits) > _PHONE_MAX_DIGITS:
            errors.append(f"phone has {len(digits)} digits (max {_PHONE_MAX_DIGITS})")
        elif digits in seen_phones:
            errors.append("duplicate phone")
        else:
            seen_phones.add(digits)

    if not address:
        errors.append("missing address")
    else:
        key = _normalize_address(address)
        if key in seen_addresses:
            errors.append("duplicate address")
        else:
            seen_addresses.add(key)

    return errors


def _validate_all(items):
    """Return list of {index, id, errors[]} for items that have any errors."""
    seen_phones = set()
    seen_addresses = set()
    issues = []
    for idx, row in enumerate(items):
        errs = _validate_row(row, seen_phones, seen_addresses)
        if errs:
            issues.append({
                "index":  idx,
                "id":     row.get("id"),
                "name":   row.get("name", ""),
                "errors": errs,
            })
    return issues


@app.route("/recipients", methods=["GET"])
def recipients_list():
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
    issues = _validate_all(doc["items"])
    return jsonify({
        "status":    "success",
        "updatedAt": doc["updatedAt"],
        "count":     len(doc["items"]),
        "items":     doc["items"],
        "issues":    issues,
    })


@app.route("/recipients/validate", methods=["POST"])
def recipients_validate():
    """Validate a candidate list without persisting it."""
    payload = request.get_json(silent=True) or {}
    items = payload.get("items")
    if not isinstance(items, list):
        return jsonify({"status": "error", "error": "items[] required"}), 400
    issues = _validate_all(items)
    return jsonify({
        "status": "success",
        "total":  len(items),
        "issues": issues,
    })


@app.route("/recipients/replace", methods=["POST"])
def recipients_replace():
    """Replace the entire recipients list. Used by the Sheets page when the
    user pastes a new CSV / saves the edit buffer."""
    payload = request.get_json(silent=True) or {}
    items_in = payload.get("items")
    if not isinstance(items_in, list):
        return jsonify({"status": "error", "error": "items[] required"}), 400

    clean = []
    for it in items_in:
        if not isinstance(it, dict):
            continue
        clean.append({
            "id":      str(it.get("id") or _new_id()),
            "name":    str(it.get("name") or "").strip(),
            "phone":   str(it.get("phone") or "").strip(),
            "address": str(it.get("address") or "").strip(),
        })

    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
        doc["items"] = clean
        _save_recipients(doc)

    issues = _validate_all(clean)
    return jsonify({
        "status":    "success",
        "count":     len(clean),
        "updatedAt": doc["updatedAt"],
        "issues":    issues,
    })


@app.route("/recipients", methods=["POST"])
def recipients_add():
    """Append one recipient row. Returns the persisted entry (with assigned id)."""
    payload = request.get_json(silent=True) or {}
    entry = {
        "id":      _new_id(),
        "name":    str(payload.get("name") or "").strip(),
        "phone":   str(payload.get("phone") or "").strip(),
        "address": str(payload.get("address") or "").strip(),
    }
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
        doc["items"].append(entry)
        _save_recipients(doc)
    return jsonify({"status": "success", "item": entry, "count": len(doc["items"])})


@app.route("/recipients/<rid>", methods=["PUT"])
def recipients_update(rid):
    payload = request.get_json(silent=True) or {}
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
        found = None
        for it in doc["items"]:
            if it.get("id") == rid:
                if "name"    in payload: it["name"]    = str(payload["name"] or "").strip()
                if "phone"   in payload: it["phone"]   = str(payload["phone"] or "").strip()
                if "address" in payload: it["address"] = str(payload["address"] or "").strip()
                found = it
                break
        if not found:
            return jsonify({"status": "error", "error": f"id {rid!r} not found"}), 404
        _save_recipients(doc)
    return jsonify({"status": "success", "item": found})


@app.route("/recipients/<rid>", methods=["DELETE"])
def recipients_delete(rid):
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
        before = len(doc["items"])
        doc["items"] = [it for it in doc["items"] if it.get("id") != rid]
        if len(doc["items"]) == before:
            return jsonify({"status": "error", "error": f"id {rid!r} not found"}), 404
        _save_recipients(doc)
    return jsonify({"status": "success", "count": len(doc["items"])})


# ---------------------------------------------------------------------------
# CSV / Excel parsing — server-side so we have one source of truth for the
# tolerated formats (header optional; columns: name, phone, address).
# ---------------------------------------------------------------------------

# Field-name aliases the importer accepts in the header row. Anything not in
# the alias map is reported as an "unknown column" rather than guessed at.
_CSV_ALIASES = {
    "name":    {"name", "full name", "recipient", "recipient name", "contact", "contact name"},
    "phone":   {"phone", "mobile", "phone number", "mobile number", "contact number", "whatsapp"},
    "address": {"address", "addr", "location", "city", "city address"},
}


def _detect_columns(header_row):
    """Map header cell positions to canonical field names. Returns
    {canonical_field: column_index} or None if no header detected."""
    norm = [(s or "").strip().lower() for s in header_row]
    mapping = {}
    for canon, aliases in _CSV_ALIASES.items():
        for i, cell in enumerate(norm):
            if cell in aliases:
                mapping[canon] = i
                break
    # Need at least 'name' OR 'phone' OR 'address' to call it a header row.
    return mapping if mapping else None


def _parse_csv_text(text):
    """Parse CSV text. Returns list of {name, phone, address}. Auto-detects
    header row using alias map; falls back to positional [name, phone, address]
    if no header is recognised."""
    if not text or not text.strip():
        return []
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return []

    header_map = _detect_columns(rows[0])
    if header_map:
        data_rows = rows[1:]
        name_i    = header_map.get("name")
        phone_i   = header_map.get("phone")
        address_i = header_map.get("address")
    else:
        # Positional: column 0=name, 1=phone, 2=address.
        data_rows = rows
        name_i, phone_i, address_i = 0, 1, 2

    out = []
    for r in data_rows:
        def cell(idx):
            return (r[idx].strip() if idx is not None and idx < len(r) and r[idx] is not None else "")
        item = {
            "name":    cell(name_i),
            "phone":   cell(phone_i),
            "address": cell(address_i),
        }
        if any(item.values()):
            out.append(item)
    return out


def _parse_xlsx_bytes(blob):
    """Parse an .xlsx workbook's first sheet. Same rules as CSV."""
    if _openpyxl is None:
        raise RuntimeError("openpyxl not installed on the server")
    wb = _openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Coerce every cell to a stripped string; treat None as empty.
        cells = [("" if c is None else str(c)).strip() for c in row]
        if any(cells):
            rows.append(cells)
    if not rows:
        return []
    header_map = _detect_columns(rows[0])
    if header_map:
        data_rows = rows[1:]
        name_i    = header_map.get("name")
        phone_i   = header_map.get("phone")
        address_i = header_map.get("address")
    else:
        data_rows = rows
        name_i, phone_i, address_i = 0, 1, 2

    out = []
    for r in data_rows:
        def cell(idx):
            return (r[idx] if idx is not None and idx < len(r) else "") or ""
        item = {
            "name":    cell(name_i).strip(),
            "phone":   cell(phone_i).strip(),
            "address": cell(address_i).strip(),
        }
        if any(item.values()):
            out.append(item)
    return out


@app.route("/recipients/import-file", methods=["POST"])
def recipients_import_file():
    """Accept a CSV or XLSX upload, parse it, and replace the recipients list.

    Query param ?dry=1 → return the parsed rows + issues WITHOUT persisting.
    Useful for the frontend to preview before committing.
    """
    if "file" not in request.files:
        return jsonify({"status": "error", "error": "no file part in request"}), 400
    f = request.files["file"]
    if not f or not (f.filename or "").strip():
        return jsonify({"status": "error", "error": "empty filename"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    blob = f.read()
    if not blob:
        return jsonify({"status": "error", "error": "file is empty"}), 400

    try:
        if ext == ".csv":
            # Decode forgivingly — most CSV exports are UTF-8 (with optional BOM).
            text = None
            for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                try:
                    text = blob.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                return jsonify({"status": "error", "error": "could not decode CSV"}), 400
            parsed = _parse_csv_text(text)
        elif ext == ".xlsx":
            if _openpyxl is None:
                return jsonify({
                    "status": "error",
                    "error":  "Excel import requires openpyxl on the server (pip install openpyxl).",
                }), 501
            parsed = _parse_xlsx_bytes(blob)
        elif ext == ".xls":
            return jsonify({
                "status": "error",
                "error":  ".xls (legacy Excel) not supported — please re-save as .xlsx or .csv.",
            }), 400
        else:
            return jsonify({
                "status": "error",
                "error":  f"unsupported extension {ext!r}; allowed: .csv, .xlsx",
            }), 400
    except Exception as e:
        return jsonify({"status": "error", "error": f"parse failed: {e}"}), 400

    if not parsed:
        return jsonify({"status": "error", "error": "no rows found in file"}), 400

    issues = _validate_all(parsed)

    dry = request.args.get("dry", "").lower() in ("1", "true", "yes")
    if not dry:
        # Assign ids and persist.
        for it in parsed:
            it["id"] = _new_id()
        with _RECIPIENTS_LOCK:
            doc = _load_recipients()
            doc["items"] = parsed
            _save_recipients(doc)

    return jsonify({
        "status":      "success",
        "filename":    f.filename,
        "extension":   ext,
        "rowsParsed":  len(parsed),
        "issues":      issues,
        "persisted":   not dry,
        "items":       parsed,
    })


# ---------------------------------------------------------------------------
# Google Sheets connection — via public CSV export URL.
#
# A Google Sheet shared as "Anyone with the link can view" exposes a CSV
# export at https://docs.google.com/spreadsheets/d/<ID>/export?format=csv&gid=<GID>
# — no OAuth required. We store the URL once, then "Force Sync" re-fetches.
# ---------------------------------------------------------------------------

_GSHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
_GSHEET_GID_RE = re.compile(r"[#?&]gid=(\d+)")


def _derive_gsheet_csv_url(url):
    """Convert a normal Google Sheets URL to its CSV-export form.
    Returns (csv_url, sheet_id, gid) or raises ValueError."""
    if not url or not isinstance(url, str):
        raise ValueError("url is required")
    m = _GSHEET_ID_RE.search(url)
    if not m:
        raise ValueError("not a Google Sheets URL (couldn't find /spreadsheets/d/<id>)")
    sheet_id = m.group(1)
    gm = _GSHEET_GID_RE.search(url)
    gid = gm.group(1) if gm else "0"
    csv_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=csv&gid={gid}"
    )
    return csv_url, sheet_id, gid


def _load_sheet_source():
    if not os.path.isfile(SHEET_SOURCE_PATH):
        return None
    try:
        with open(SHEET_SOURCE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        return data
    except (OSError, json.JSONDecodeError):
        return None


def _save_sheet_source(data):
    os.makedirs(os.path.dirname(SHEET_SOURCE_PATH), exist_ok=True)
    tmp = SHEET_SOURCE_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")
    os.replace(tmp, SHEET_SOURCE_PATH)


def _delete_sheet_source():
    if os.path.isfile(SHEET_SOURCE_PATH):
        try:
            os.remove(SHEET_SOURCE_PATH)
        except OSError:
            pass


_ALLOWED_GSHEET_HOSTS = {
    "docs.google.com",
    "googleusercontent.com",
    "www.googleusercontent.com",
}


def _fetch_gsheet_csv(csv_url):
    """Download the CSV. Raises with an operator-readable message on failure.

    SSRF hardening:
      - Connect + read split into a (5, 20) timeout pair so a hung host
        can't lock the request thread for the full read window.
      - `allow_redirects=False` plus an explicit one-hop follow that
        validates the next URL is still on a Google host. Prevents the
        CSV endpoint from being used as an open-redirect into internal
        services.
    """
    if _requests is None:
        raise RuntimeError("the 'requests' package is not installed on the server")
    try:
        resp = _requests.get(csv_url, timeout=(5, 20), allow_redirects=False)
        # Manual one-hop follow with host whitelist. Google's CSV endpoint
        # often 30x to a googleusercontent.com signed URL — allow that, but
        # refuse anything else.
        if 300 <= resp.status_code < 400:
            next_url = resp.headers.get("Location") or ""
            try:
                from urllib.parse import urlparse
                host = urlparse(next_url).hostname or ""
            except ValueError:
                host = ""
            if not any(host == h or host.endswith("." + h) for h in _ALLOWED_GSHEET_HOSTS):
                raise RuntimeError(
                    f"refused redirect to non-Google host: {host or '<unknown>'}"
                )
            resp = _requests.get(next_url, timeout=(5, 20), allow_redirects=False)
    except _requests.exceptions.RequestException as e:
        raise RuntimeError(f"network error: {e}") from e
    if resp.status_code == 401 or resp.status_code == 403:
        raise RuntimeError(
            "sheet is not publicly readable — share it as "
            "'Anyone with the link can view' first"
        )
    if resp.status_code >= 400:
        raise RuntimeError(f"Google responded HTTP {resp.status_code}")
    ctype = (resp.headers.get("Content-Type") or "").lower()
    text = resp.text
    if "text/html" in ctype and "<html" in text.lower()[:200]:
        raise RuntimeError("Google returned HTML, not CSV — sheet likely private")
    return text


@app.route("/recipients/sheet-source", methods=["GET"])
def sheet_source_get():
    src = _load_sheet_source()
    return jsonify({"status": "success", "source": src})


@app.route("/recipients/sheet-source", methods=["DELETE"])
def sheet_source_delete():
    _delete_sheet_source()
    return jsonify({"status": "success"})


@app.route("/recipients/connect-google-sheet", methods=["POST"])
def connect_google_sheet():
    """Connect (and immediately sync) a public Google Sheet."""
    payload = request.get_json(silent=True) or {}
    url = (payload.get("url") or "").strip()
    try:
        csv_url, sheet_id, gid = _derive_gsheet_csv_url(url)
    except ValueError as e:
        return jsonify({"status": "error", "error": str(e)}), 400

    try:
        text = _fetch_gsheet_csv(csv_url)
    except RuntimeError as e:
        return jsonify({"status": "error", "error": str(e)}), 502

    try:
        parsed = _parse_csv_text(text)
    except Exception as e:
        return jsonify({"status": "error", "error": f"parse failed: {e}"}), 400
    if not parsed:
        return jsonify({"status": "error", "error": "sheet produced 0 rows"}), 400

    for it in parsed:
        it["id"] = _new_id()
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
        doc["items"] = parsed
        _save_recipients(doc)

    now = int(time.time() * 1000)
    source = {
        "type":     "google-sheet",
        "url":      url,
        "sheetId":  sheet_id,
        "gid":      gid,
        "csvUrl":   csv_url,
        "lastSync": now,
        "rowCount": len(parsed),
        "status":   "ok",
    }
    _save_sheet_source(source)

    issues = _validate_all(parsed)
    return jsonify({
        "status":     "success",
        "source":     source,
        "rowsParsed": len(parsed),
        "issues":     issues,
    })


@app.route("/recipients/sync-google-sheet", methods=["POST"])
def sync_google_sheet():
    """Re-fetch the currently connected sheet."""
    src = _load_sheet_source()
    if not src or src.get("type") != "google-sheet":
        return jsonify({"status": "error", "error": "no Google Sheet connected"}), 400
    csv_url = src.get("csvUrl")
    if not csv_url:
        return jsonify({"status": "error", "error": "stored source is missing csvUrl"}), 500

    try:
        text = _fetch_gsheet_csv(csv_url)
    except RuntimeError as e:
        # Record the failure on the source so the UI can surface it.
        src["status"] = "error"
        src["lastError"] = str(e)
        src["lastSync"] = int(time.time() * 1000)
        _save_sheet_source(src)
        return jsonify({"status": "error", "error": str(e), "source": src}), 502

    try:
        parsed = _parse_csv_text(text)
    except Exception as e:
        return jsonify({"status": "error", "error": f"parse failed: {e}"}), 400

    for it in parsed:
        it["id"] = _new_id()
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()
        doc["items"] = parsed
        _save_recipients(doc)

    src["status"]    = "ok"
    src["lastSync"]  = int(time.time() * 1000)
    src["rowCount"]  = len(parsed)
    src.pop("lastError", None)
    _save_sheet_source(src)

    issues = _validate_all(parsed)
    return jsonify({
        "status":     "success",
        "source":     src,
        "rowsParsed": len(parsed),
        "issues":     issues,
    })


@app.route("/recipients/export.csv", methods=["GET"])
def recipients_export_csv():
    """Download recipients as a CSV for the 'Open Sheet' button."""
    with _RECIPIENTS_LOCK:
        doc = _load_recipients()

    def _csv_cell(s):
        s = "" if s is None else str(s)
        if any(c in s for c in [",", "\"", "\n", "\r"]):
            return '"' + s.replace('"', '""') + '"'
        return s

    lines = ["name,phone,address"]
    for it in doc["items"]:
        lines.append(",".join([
            _csv_cell(it.get("name")),
            _csv_cell(it.get("phone")),
            _csv_cell(it.get("address")),
        ]))
    body = "\n".join(lines) + "\n"
    return (
        body,
        200,
        {
            "Content-Type":        "text/csv; charset=utf-8",
            "Content-Disposition": 'attachment; filename="recipients.csv"',
        },
    )


# ---------------------------------------------------------------------------
# Asynchronous generation jobs — produces live progress 1/N, 2/N, ...
# ---------------------------------------------------------------------------

_JOBS = {}              # job_id -> dict
_JOBS_LOCK = threading.Lock()
_MAX_JOBS = 32          # cap the in-memory job table


def _job_file_path(job_id):
    return os.path.join(JOBS_DIR, f"{job_id}.json")


def _persist_job(job):
    try:
        path = _job_file_path(job.get("id", ""))
        with open(path, "w", encoding="utf-8") as f:
            json.dump(job, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _load_job(job_id):
    path = _job_file_path(job_id)
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _run_job(job_id, recipients, kind="all"):
    """Wrapper: catches any unhandled exception and lands the job in
    state='error' so the worker thread never dies silently."""
    try:
        _run_job_inner(job_id, recipients, kind)
    except Exception as e:                                  # noqa: BLE001
        # Last-ditch safety net. Logs to the Flask terminal AND moves the
        # job out of 'running' so the eviction LRU + the frontend poller
        # both see a terminal state.
        import traceback
        traceback.print_exc()
        with _JOBS_LOCK:
            if job_id in _JOBS:
                _JOBS[job_id].update({
                    "state":      "error",
                    "error":      f"worker exception: {type(e).__name__}: {e}",
                    "finishedAt": int(time.time() * 1000),
                    "updatedAt":  int(time.time() * 1000),
                    "current":    None,
                })
                _persist_job(_JOBS[job_id])
        print(
            f"[/generate-jobs {job_id}] FAILED kind={kind.upper()} -- {type(e).__name__}: {e}",
            flush=True,
        )


def _run_job_inner(job_id, recipients, kind="all"):
    """Worker thread: runs the requested generators for each recipient
    sequentially and records results as it goes.

    `kind` controls which generators run:
      "all"    — image + video (legacy default)
      "images" — image only
      "videos" — video only
    Sequential by design — image_generator and video_generator both invoke
    ffmpeg/Pillow + read the same template, so parallelising would mostly
    stress the disk."""
    def _set(**kwargs):
        with _JOBS_LOCK:
            if job_id in _JOBS:
                _JOBS[job_id].update(kwargs)
                _JOBS[job_id]["updatedAt"] = int(time.time() * 1000)
                _persist_job(_JOBS[job_id])

    _set(state="running", startedAt=int(time.time() * 1000))

    results = []
    for i, r in enumerate(recipients):
        # Pause/cancel gate. The worker keeps the current row atomic (so we
        # don't ship half-rendered recipients), then idles here until the
        # client either resumes or cancels.
        #
        # `pauseEffective` tells the UI the worker has actually reached this
        # gate — i.e. nothing is rendering right now. Between a user clicking
        # Pause and the gate being reached, the row in flight is still
        # finishing; the UI shows "Pausing…" during that window.
        while True:
            with _JOBS_LOCK:
                j = _JOBS.get(job_id, {})
                if j.get("cancelled"):
                    _set(state="cancelled", finishedAt=int(time.time() * 1000))
                    return
                if not j.get("paused"):
                    if j.get("pauseEffective"):
                        _JOBS[job_id]["pauseEffective"] = False
                    _JOBS[job_id]["current"] = {
                        "index":   i,
                        "name":    r.get("name", ""),
                        "address": r.get("address", ""),
                        "phone":   r.get("phone", ""),
                    }
                    _JOBS[job_id]["updatedAt"] = int(time.time() * 1000)
                    break
                # Paused — mark that the worker is actually idle so the UI
                # can distinguish "pausing" (current row still rendering)
                # from "paused" (queue truly halted).
                if not j.get("pauseEffective"):
                    _JOBS[job_id]["pauseEffective"] = True
                    _JOBS[job_id]["updatedAt"]      = int(time.time() * 1000)
            time.sleep(0.3)

        address = (r.get("address") or "").strip()
        phone   = (r.get("phone") or "").strip()
        name    = (r.get("name") or "").strip()

        if not address or not phone:
            results.append({
                "index": i, "id": r.get("id"), "name": name,
                "address": address, "phone": phone,
                "ok": False, "error": "address and phone are required",
            })
        else:
            # MEDIA TYPE dispatch — strictly one pipeline per kind. There is
            # no shared loop, no Promise.all, no "run both then filter."
            # Each kind has its OWN dedicated pipeline function.
            print(
                f"[/generate-jobs {job_id} row {i + 1}/{len(recipients)}] "
                f"MEDIA TYPE={kind.upper()} name={name!r} addr={address!r}",
                flush=True,
            )
            image = None
            video = None
            if kind == "images":
                image = _run_image_pipeline(address, phone, name)
            elif kind == "videos":
                video = _run_video_pipeline(address, phone, name)
            elif kind == "all":
                # Legacy kind — explicit AND of the two named pipelines.
                # Used only when a caller explicitly opts in via kind='all'.
                image = _run_image_pipeline(address, phone, name)
                video = _run_video_pipeline(address, phone, name)
            row_ok = (image["ok"] if image else True) and (video["ok"] if video else True)

            def _tail(s, n=400):
                s = (s or "").strip()
                return s[-n:] if len(s) > n else s

            row = {
                "index":   i,
                "id":      r.get("id"),
                "name":    name,
                "address": address,
                "phone":   phone,
                "ok":      row_ok,
            }
            if image is not None:
                row["image"] = {
                    "ok":     image["ok"],
                    "path":   image["path"],
                    "stderr": _tail(image["stderr"]) if not image["ok"] else "",
                }
            if video is not None:
                row["video"] = {
                    "ok":     video["ok"],
                    "path":   video["path"],
                    "stderr": _tail(video["stderr"]) if not video["ok"] else "",
                }
            results.append(row)

        with _JOBS_LOCK:
            if job_id in _JOBS:
                _JOBS[job_id]["progress"] = i + 1
                _JOBS[job_id]["results"]  = list(results)

    succeeded = sum(1 for x in results if x["ok"])
    # Count which kinds of files were actually produced — make pipeline
    # isolation visible in the logs.
    image_count = sum(1 for r in results if (r.get("image") or {}).get("ok"))
    video_count = sum(1 for r in results if (r.get("video") or {}).get("ok"))
    print(
        f"[/generate-jobs {job_id}] FINISHED kind={kind.upper()} "
        f"succeeded={succeeded}/{len(results)} "
        f"produced: {image_count} image(s), {video_count} video(s)",
        flush=True,
    )
    final_state = "done"
    _set(
        state=final_state,
        finishedAt=int(time.time() * 1000),
        succeeded=succeeded,
        failed=len(results) - succeeded,
        current=None,
    )


def _evict_oldest_jobs():
    with _JOBS_LOCK:
        if len(_JOBS) <= _MAX_JOBS:
            return
        # Drop the oldest finished/cancelled/errored job first. 'error' is
        # a terminal state too — the worker wrapper sets it when an
        # unhandled exception escapes the loop.
        finished = [j for j in _JOBS.values() if j.get("state") in ("done", "cancelled", "error")]
        finished.sort(key=lambda j: j.get("finishedAt") or j.get("startedAt") or 0)
        while len(_JOBS) > _MAX_JOBS and finished:
            victim = finished.pop(0)
            _JOBS.pop(victim["id"], None)


@app.route("/generate-jobs", methods=["POST"])
def generate_jobs_create():
    """Kick off an async generation run. Returns a job_id the client polls."""
    payload = request.get_json(silent=True) or {}
    recipients = payload.get("recipients")
    if not isinstance(recipients, list) or not recipients:
        return jsonify({"status": "error", "error": "recipients[] required"}), 400

    kind = (payload.get("kind") or "all").strip().lower()
    if kind not in ("all", "images", "videos"):
        return jsonify({"status": "error", "error": "kind must be 'all', 'images', or 'videos'"}), 400

    job_id = uuid.uuid4().hex[:12]
    # Top-of-job log so the user can confirm in the Flask terminal exactly
    # which pipeline was kicked off. If you click "Generate All Images" you
    # should see kind=images here and only image-pipeline rows below it.
    print(
        f"[/generate-jobs {job_id}] STARTED kind={kind.upper()} "
        f"count={len(recipients)} "
        f"-> will run: image={'YES' if kind in ('all','images') else 'NO'} "
        f"video={'YES' if kind in ('all','videos') else 'NO'}",
        flush=True,
    )
    job = {
        "id":             job_id,
        "kind":           kind,
        "state":          "pending",
        "total":          len(recipients),
        "progress":       0,
        "current":        None,
        "results":        [],
        "succeeded":      0,
        "failed":         0,
        "cancelled":      False,
        "paused":         False,
        # Worker sets True once it actually reaches the pause gate. The UI
        # uses this to show "Pausing…" vs "Paused".
        "pauseEffective": False,
        "createdAt":      int(time.time() * 1000),
        "startedAt":      None,
        "finishedAt":     None,
        "updatedAt":      int(time.time() * 1000),
    }
    with _JOBS_LOCK:
        _JOBS[job_id] = job
        _persist_job(job)
    _evict_oldest_jobs()

    t = threading.Thread(target=_run_job, args=(job_id, recipients, kind), daemon=True)
    t.start()

    return jsonify({"status": "success", "job_id": job_id, "total": len(recipients), "kind": kind})


@app.route("/generate-jobs/<job_id>", methods=["GET"])
def generate_jobs_status(job_id):
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            job = _load_job(job_id)
            if job:
                _JOBS[job_id] = job
        if not job:
            return jsonify({"status": "error", "error": "job not found"}), 404
        # Return a shallow copy without the internal cancelled flag.
        body = {k: v for k, v in job.items() if k != "cancelled"}
    body["status"] = "success"
    return jsonify(body)


@app.route("/generate-jobs/<job_id>/cancel", methods=["POST"])
def generate_jobs_cancel(job_id):
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            job = _load_job(job_id)
            if job:
                _JOBS[job_id] = job
        if not job:
            return jsonify({"status": "error", "error": "job not found"}), 404
        if job["state"] not in ("pending", "running"):
            return jsonify({"status": "error", "error": f"job is {job['state']}"}), 400
        job["cancelled"] = True
        job["updatedAt"] = int(time.time() * 1000)
        _persist_job(job)
    return jsonify({"status": "success", "job_id": job_id})


@app.route("/generate-jobs/<job_id>/pause", methods=["POST"])
def generate_jobs_pause(job_id):
    """Pause an in-flight generation job. The worker thread keeps the
    current row in-flight (to avoid partial files), then idles until
    /resume is called or the job is cancelled."""
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            job = _load_job(job_id)
            if job:
                _JOBS[job_id] = job
        if not job:
            return jsonify({"status": "error", "error": "job not found"}), 404
        if job["state"] not in ("pending", "running"):
            return jsonify({
                "status": "error",
                "error":  f"cannot pause a {job['state']} job",
            }), 400
        job["paused"]    = True
        job["updatedAt"] = int(time.time() * 1000)
        _persist_job(job)
    return jsonify({"status": "success", "job_id": job_id})


@app.route("/generate-jobs/<job_id>/resume", methods=["POST"])
def generate_jobs_resume(job_id):
    """Resume a paused generation job. Like /pause, refuses to act on
    jobs that have already reached a terminal state — silently flipping
    `paused=false` on a done/cancelled/error job would just leave a stale
    flag in the response."""
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            job = _load_job(job_id)
            if job:
                _JOBS[job_id] = job
        if not job:
            return jsonify({"status": "error", "error": "job not found"}), 404
        if job["state"] not in ("pending", "running"):
            return jsonify({
                "status": "error",
                "error":  f"cannot resume a {job['state']} job",
            }), 400
        job["paused"]         = False
        # `pauseEffective` clears itself when the worker passes the gate;
        # but flip it eagerly so the UI doesn't briefly flash "Paused"
        # after a resume click.
        job["pauseEffective"] = False
        job["updatedAt"]      = int(time.time() * 1000)
        _persist_job(job)
    return jsonify({"status": "success", "job_id": job_id})


# ---------------------------------------------------------------------------
# Bulk generation across a recipient list (legacy sync endpoint — kept for
# curl / scripting; the UI now uses /generate-jobs for live progress)
# ---------------------------------------------------------------------------

@app.route("/generate-videos", methods=["POST"])
def generate_videos():
    """Legacy synchronous bulk endpoint. The UI no longer calls this (it
    uses /generate-jobs for async + kind-aware runs), but the route is
    kept for scripted/curl use. To preserve pipeline isolation, callers
    MUST now pass an explicit `kind` of either "image" or "video" — we
    no longer silently run both."""
    payload = request.get_json(silent=True) or {}
    recipients = payload.get("recipients")
    kind = (payload.get("kind") or "").strip().lower()

    if not isinstance(recipients, list) or not recipients:
        return jsonify({
            "status": "error",
            "error": "recipients[] is required (non-empty list)",
        }), 400
    if kind not in ("image", "video"):
        return jsonify({
            "status": "error",
            "error": "kind is required and must be 'image' or 'video' — "
                     "this endpoint no longer runs both pipelines together",
        }), 400

    print(f"[/generate-videos] kind={kind.upper()} count={len(recipients)}", flush=True)

    results = []
    succeeded = 0

    for i, r in enumerate(recipients):
        if not isinstance(r, dict):
            results.append({
                "index": i,
                "ok": False,
                "error": "row is not an object",
            })
            continue

        address = str(r.get("address") or "").strip()
        phone   = str(r.get("phone") or "").strip()
        name    = str(r.get("name") or "").strip()

        if not address or not phone:
            results.append({
                "index": i, "name": name, "address": address, "phone": phone,
                "ok": False, "error": "address and phone are required",
            })
            continue

        print(
            f"[/generate-videos {i + 1}/{len(recipients)}] kind={kind} "
            f"name={name!r} addr={address!r} phone={phone!r}",
            flush=True,
        )

        image = _run_image_pipeline(address, phone, name) if kind == "image" else None
        video = _run_video_pipeline(address, phone, name) if kind == "video" else None
        row_ok = (image["ok"] if image else True) and (video["ok"] if video else True)
        if row_ok:
            succeeded += 1

        # Trim noisy stderr to the tail when present — useful for UI display
        # without flooding the JSON response.
        def _tail(s, n=400):
            s = (s or "").strip()
            return s[-n:] if len(s) > n else s

        row = {
            "index":   i,
            "name":    name,
            "address": address,
            "phone":   phone,
            "ok":      row_ok,
        }
        if image is not None:
            row["image"] = {
                "ok":     image["ok"],
                "path":   image["path"],
                "stderr": _tail(image["stderr"]) if not image["ok"] else "",
            }
        if video is not None:
            row["video"] = {
                "ok":     video["ok"],
                "path":   video["path"],
                "stderr": _tail(video["stderr"]) if not video["ok"] else "",
            }
        results.append(row)

    if succeeded == len(results):
        status = "success"
    elif succeeded > 0:
        status = "partial"
    else:
        status = "error"

    return jsonify({
        "status":    status,
        "total":     len(results),
        "succeeded": succeeded,
        "failed":    len(results) - succeeded,
        "results":   results,
    }), (200 if succeeded > 0 else 500)


# ---------------------------------------------------------------------------
# Filesystem-driven listing of generated outputs + static serving
# ---------------------------------------------------------------------------

def _scan_output_dir(dir_path, allowed_exts):
    """Return { stem (without ext): {filename, size, mtime} } for non-empty files."""
    result = {}
    if not os.path.isdir(dir_path):
        return result
    try:
        entries = os.listdir(dir_path)
    except OSError:
        return result
    for entry in entries:
        full = os.path.join(dir_path, entry)
        if not os.path.isfile(full):
            continue
        stem, ext = os.path.splitext(entry)
        if ext.lower() not in allowed_exts:
            continue
        try:
            size = os.path.getsize(full)
            mtime = os.path.getmtime(full)
        except OSError:
            continue
        if size <= 0:
            continue  # skip empty / partial writes
        # If two files share a stem with different extensions, the first wins —
        # generators only emit one extension per kind so this is effectively
        # never a real conflict.
        result.setdefault(stem, {
            "filename": entry,
            "size":     size,
            "mtime":    mtime,
        })
    return result


def _stem_to_display_name(stem):
    """Convert sanitised filename stem back to a readable label.

    Generators sanitise via underscore-substitution: "SKF Colony Pune"
    becomes "SKF_Colony_Pune". This reverses that for display.
    """
    return stem.replace("_", " ")


def _list_generated_items():
    """Build the canonical list of generated entries from disk.

    Shared by /list-generated, /delivery-status, and /dashboard-stats so
    every page sees the same data with the same status rules.

    Status:
      * "Queued" — any media file exists for this stem (ready to be
        delivered). Images and videos are now generated by independent
        pipelines, so a stem with only one half is NOT a failure — the
        user explicitly chose to render that kind on its own.
    """
    videos = _scan_output_dir(OUTPUT_VIDEOS, _VIDEO_OUT_EXTS)
    images = _scan_output_dir(OUTPUT_IMAGES, _IMAGE_OUT_EXTS)

    stems = set(videos.keys()) | set(images.keys())
    items = []
    for stem in stems:
        v = videos.get(stem)
        i = images.get(stem)
        mtimes = [m["mtime"] for m in (v, i) if m]
        if not mtimes:
            continue

        item = {
            "id":        stem,
            "name":      _stem_to_display_name(stem),
            "video":     None,
            "image":     None,
            "status":    "Queued",
            "createdAt": int(max(mtimes) * 1000),
        }
        if v:
            item["video"] = {
                "filename": v["filename"],
                "url":      _session.make_signed_url(f"/files/videos/{v['filename']}"),
                "size":     v["size"],
            }
        if i:
            item["image"] = {
                "filename": i["filename"],
                "url":      _session.make_signed_url(f"/files/images/{i['filename']}"),
                "size":     i["size"],
            }
        items.append(item)

    items.sort(key=lambda x: x["createdAt"], reverse=True)
    return items


@app.route("/list-generated", methods=["GET"])
def list_generated():
    items = _list_generated_items()
    return jsonify({"status": "success", "count": len(items), "items": items})


def _is_safe_stem(stem):
    """Path-traversal guard for filename stems."""
    if not stem:
        return False
    return not (
        "/" in stem
        or "\\" in stem
        or ".." in stem
        or stem.startswith(".")
    )


@app.route("/list-generated/delete", methods=["POST"])
def list_generated_delete():
    """Remove generated outputs.

    Per-asset (preferred — Generated Media uses this):
        { items: [{ stem: "SKF_Colony_Pune", kind: "image" }, ...] }
      Each entry deletes ONLY the named half. The matching delivery
      record is dropped only when no media remains for that stem.

    Legacy stem-level (kept for older scripts / curl):
        { ids: ["SKF_Colony_Pune", ...] }
      Each id deletes both halves (image + video) for the stem.

    Path-traversal guarded. Returns per-asset reporting so the UI can show
    "deleted 3 images, 0 videos failed".
    """
    payload = request.get_json(silent=True) or {}
    items = payload.get("items")
    legacy_ids = payload.get("ids")

    if not isinstance(items, list):
        items = None
    if not isinstance(legacy_ids, list):
        legacy_ids = None
    if not items and not legacy_ids:
        return jsonify({"status": "error", "error": "items[] or ids[] required"}), 400

    # Normalise both shapes to a per-asset todo list: [(stem, kind), ...]
    targets = []
    if items:
        for entry in items:
            if not isinstance(entry, dict):
                continue
            stem = str(entry.get("stem") or "").strip()
            kind = str(entry.get("kind") or "").strip().lower()
            if kind in ("image", "video"):
                targets.append((stem, kind))
    if legacy_ids:
        for raw_id in legacy_ids:
            stem = str(raw_id or "").strip()
            # Legacy: expand to both halves.
            targets.append((stem, "image"))
            targets.append((stem, "video"))

    if not targets:
        return jsonify({"status": "error", "error": "no valid items"}), 400

    deleted = []   # [{stem, kind, filename}]
    failed  = []   # [{stem, kind, error}]
    touched_stems = set()

    for stem, kind in targets:
        if not _is_safe_stem(stem):
            failed.append({"stem": stem, "kind": kind, "error": "invalid stem"})
            continue

        if kind == "image":
            dir_path, exts = OUTPUT_IMAGES, _IMAGE_OUT_EXTS
        else:
            dir_path, exts = OUTPUT_VIDEOS, _VIDEO_OUT_EXTS

        removed_for_this = False
        for ext in exts:
            fpath = os.path.join(dir_path, stem + ext)
            if os.path.isfile(fpath):
                try:
                    os.remove(fpath)
                    deleted.append({
                        "stem":     stem,
                        "kind":     kind,
                        "filename": os.path.basename(fpath),
                    })
                    removed_for_this = True
                    touched_stems.add(stem)
                except OSError as e:
                    failed.append({
                        "stem":  stem,
                        "kind":  kind,
                        "error": f"remove {fpath}: {e}",
                    })
        if not removed_for_this:
            failed.append({
                "stem":  stem,
                "kind":  kind,
                "error": f"no {kind} file found",
            })

    # Delivery records key off the stem (recipient), not the asset kind.
    # Only drop a delivery row when BOTH halves are gone — otherwise the
    # recipient still has media on disk and the row stays meaningful.
    removed_dlv = 0
    if touched_stems:
        with _DELIVERIES_LOCK:
            doc = _load_deliveries()
            kept = []
            for d in doc["items"]:
                s = d.get("stem")
                if s in touched_stems:
                    # Check filesystem — drop only if nothing left.
                    img_left = any(
                        os.path.isfile(os.path.join(OUTPUT_IMAGES, s + e))
                        for e in _IMAGE_OUT_EXTS
                    )
                    vid_left = any(
                        os.path.isfile(os.path.join(OUTPUT_VIDEOS, s + e))
                        for e in _VIDEO_OUT_EXTS
                    )
                    if img_left or vid_left:
                        kept.append(d)
                    else:
                        removed_dlv += 1
                else:
                    kept.append(d)
            if removed_dlv:
                doc["items"] = kept
                _save_deliveries(doc)

    print(
        f"[/list-generated/delete] removed {len(deleted)} asset(s) "
        f"({sum(1 for d in deleted if d['kind'] == 'image')} image, "
        f"{sum(1 for d in deleted if d['kind'] == 'video')} video), "
        f"{len(failed)} failed, dropped {removed_dlv} delivery row(s)",
        flush=True,
    )

    return jsonify({
        "status":     "success" if not failed else ("partial" if deleted else "error"),
        "deleted":    deleted,
        "failed":     failed,
        "count":      len(deleted),
        "deliveries": removed_dlv,
    })


@app.route("/list-generated/wipe", methods=["POST"])
def list_generated_wipe():
    """Bulk-delete every file of the requested kind(s) so the gallery can
    be cleared in one shot. Lets a user prove pipeline isolation by
    starting from an empty slate.

    Body: { kinds: ["image"] | ["video"] | ["image","video"] }

    Always requires explicit `kinds` — never silently deletes everything.
    """
    payload = request.get_json(silent=True) or {}
    kinds = payload.get("kinds")
    if not isinstance(kinds, list) or not kinds:
        return jsonify({"status": "error", "error": "kinds[] required"}), 400
    kinds = [str(k).strip().lower() for k in kinds]
    invalid = [k for k in kinds if k not in ("image", "video")]
    if invalid:
        return jsonify({
            "status": "error",
            "error":  f"invalid kind(s): {invalid}; allowed: image, video",
        }), 400

    deleted = []
    failed  = []

    for kind in kinds:
        dir_path, exts = (
            (OUTPUT_IMAGES, _IMAGE_OUT_EXTS) if kind == "image"
            else (OUTPUT_VIDEOS, _VIDEO_OUT_EXTS)
        )
        try:
            files = os.listdir(dir_path)
        except FileNotFoundError:
            continue
        for fn in files:
            stem, ext = os.path.splitext(fn)
            if ext.lower() not in exts:
                continue
            fpath = os.path.join(dir_path, fn)
            try:
                os.remove(fpath)
                deleted.append({"stem": stem, "kind": kind, "filename": fn})
            except OSError as e:
                failed.append({"filename": fn, "kind": kind, "error": str(e)})

    # Clean dangling delivery records — any stem with no remaining media.
    removed_dlv = 0
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        kept = []
        for d in doc["items"]:
            s = d.get("stem")
            img_left = any(
                os.path.isfile(os.path.join(OUTPUT_IMAGES, s + e))
                for e in _IMAGE_OUT_EXTS
            )
            vid_left = any(
                os.path.isfile(os.path.join(OUTPUT_VIDEOS, s + e))
                for e in _VIDEO_OUT_EXTS
            )
            if img_left or vid_left:
                kept.append(d)
            else:
                removed_dlv += 1
        if removed_dlv:
            doc["items"] = kept
            _save_deliveries(doc)

    print(
        f"[/list-generated/wipe] kinds={kinds} removed {len(deleted)} file(s) "
        f"({sum(1 for d in deleted if d['kind'] == 'image')} image, "
        f"{sum(1 for d in deleted if d['kind'] == 'video')} video), "
        f"dropped {removed_dlv} delivery row(s), failed={len(failed)}",
        flush=True,
    )

    return jsonify({
        "status":     "success" if not failed else "partial",
        "deleted":    deleted,
        "failed":     failed,
        "count":      len(deleted),
        "deliveries": removed_dlv,
    })


@app.route("/delivery-status", methods=["GET"])
def delivery_status():
    """Cross-joined view of filesystem outputs + delivery records.

    For each generated stem:
      * if a delivery record exists → use its status/recipient/attempts/etc.
      * if not → fall back to filesystem-derived status ("Queued" if both
        halves present, "Failed" if only one).
    """
    merged, counts = _materialise_delivery_view()
    return jsonify({
        "status":  "success",
        "count":   len(merged),
        "counts":  counts,
        "items":   merged,
        "worker":  _worker_status(),
    })


def _count_templates():
    """Count uploaded/master template files in templates/. Excludes the
    Mastermind logo asset which lives in the same folder."""
    if not os.path.isdir(TEMPLATES_DIR):
        return 0
    n = 0
    for entry in os.listdir(TEMPLATES_DIR):
        full = os.path.join(TEMPLATES_DIR, entry)
        if not os.path.isfile(full):
            continue
        if "logo" in entry.lower():
            continue
        ext = os.path.splitext(entry)[1].lower()
        if ext in (_VIDEO_OUT_EXTS | _IMAGE_OUT_EXTS):
            try:
                if os.path.getsize(full) > 0:
                    n += 1
            except OSError:
                continue
    return n


@app.route("/dashboard-stats", methods=["GET"])
def dashboard_stats():
    """Headline numbers for the Dashboard KPI row, including real
    delivery counts when delivery records exist."""
    items_view, counts_view = _materialise_delivery_view()

    total_videos     = sum(1 for it in items_view if it.get("video"))
    total_images     = sum(1 for it in items_view if it.get("image"))
    total_recipients = len(items_view)
    delivered        = counts_view.get("Delivered", 0)
    failed           = counts_view.get("Failed", 0)

    # Break the delivered total down by which media kind was actually sent.
    # A row may carry both an image and a video — counting both honors the
    # "Media Sent" KPI's split helper "X videos sent · Y images sent."
    videos_sent = sum(
        1 for it in items_view if it.get("status") == "Delivered" and it.get("video")
    )
    images_sent = sum(
        1 for it in items_view if it.get("status") == "Delivered" and it.get("image")
    )

    now = datetime.now()
    start_of_today_ms = int(datetime(now.year, now.month, now.day).timestamp() * 1000)
    generated_today = sum(
        1 for it in items_view if (it.get("createdAt") or 0) >= start_of_today_ms
    )

    return jsonify({
        "status": "success",
        "stats": {
            "totalVideos":     total_videos,
            "totalImages":     total_images,
            "totalTemplates":  _count_templates(),
            "totalRecipients": total_recipients,
            "generatedToday":  generated_today,
            "deliveredCount":  delivered,
            "videosSent":      videos_sent,
            "imagesSent":      images_sent,
            "failedCount":     failed,
        },
        "latest": items_view[:6],
    })


def _verify_file_signature(rel_path):
    """Reject the request unless the URL carries a valid `?exp=&sig=` from
    `_session.make_signed_url`. Returns a 403 Flask response if the
    signature is missing / tampered / expired, or None to proceed.

    Templates fetched directly by the operator UI (Templates page preview)
    skip this check because the React app can attach a Bearer token via
    fetch — but `<video src>` / `<img src>` in the gallery cannot, which
    is the whole reason signed URLs exist."""
    exp = request.args.get("exp")
    sig = request.args.get("sig")
    if _session.verify_signed_path(rel_path, exp, sig):
        return None
    return jsonify({
        "status": "error",
        "error":  "invalid or expired media URL",
    }), 403


@app.route("/files/videos/<path:filename>", methods=["GET"])
def serve_video_file(filename):
    # Signed-URL gate. send_from_directory then handles path traversal.
    blocked = _verify_file_signature(f"/files/videos/{filename}")
    if blocked:
        return blocked
    return send_from_directory(OUTPUT_VIDEOS, filename, conditional=True)


@app.route("/files/images/<path:filename>", methods=["GET"])
def serve_image_file(filename):
    blocked = _verify_file_signature(f"/files/images/{filename}")
    if blocked:
        return blocked
    return send_from_directory(OUTPUT_IMAGES, filename, conditional=True)


@app.route("/files/templates/<path:filename>", methods=["GET"])
def serve_template_file(filename):
    """Templates are gated by the regular Bearer-token middleware (this
    route is NOT in _PUBLIC_PREFIXES), so signed URLs aren't needed —
    only an authenticated operator can fetch them."""
    return send_from_directory(TEMPLATES_DIR, filename, conditional=True)


@app.route("/current-template", methods=["GET"])
def current_template():
    """Return the template currently pointed at by config/settings.json for
    the requested kind. If file missing, restore from base64 in settings.json.

    Query: ?kind=video | ?kind=image
    """
    kind = (request.args.get("kind") or "").strip().lower()
    if kind not in ("video", "image"):
        return jsonify({
            "status": "error",
            "error": "kind must be 'video' or 'image'",
        }), 400

    try:
        settings = _load_settings()
    except (OSError, json.JSONDecodeError) as e:
        return jsonify({
            "status": "error",
            "error": f"cannot read settings.json: {e}",
        }), 500

    rel_path = settings.get(f"template_{kind}")
    if not rel_path:
        return jsonify({
            "status": "error",
            "error": f"no template_{kind} configured",
        }), 404

    abs_path = os.path.join(PROJECT_ROOT, rel_path)
    
    # If file doesn't exist, try to restore from base64 in settings.json
    if not os.path.isfile(abs_path):
        data_key = f"template_{kind}_data"
        if data_key in settings:
            try:
                import base64
                file_data = base64.b64decode(settings[data_key])
                os.makedirs(TEMPLATES_DIR, exist_ok=True)
                with open(abs_path, 'wb') as fp:
                    fp.write(file_data)
                print(
                    f"[/current-template] restored {kind} from base64 ({len(file_data)} bytes)",
                    flush=True,
                )
            except Exception as e:
                print(f"[/current-template] restoration failed: {e}", flush=True)
                return jsonify({
                    "status": "error",
                    "error": "template missing and restoration failed",
                }), 404
        else:
            return jsonify({
                "status": "error",
                "error": f"template file missing on disk: {rel_path}",
            }), 404

    try:
        size  = os.path.getsize(abs_path)
        mtime = os.path.getmtime(abs_path)
    except OSError as e:
        return jsonify({"status": "error", "error": str(e)}), 500
    if size <= 0:
        return jsonify({"status": "error", "error": "template file is empty"}), 404

    filename = os.path.basename(rel_path)
    return jsonify({
        "status":   "success",
        "type":     kind,
        "filename": filename,
        "url":      f"/files/templates/{filename}",
        "size":     size,
        "mtime":    int(mtime * 1000),
        "path":     rel_path,
    })


# ===========================================================================
# Delivery subsystem — WhatsApp send orchestration
# ===========================================================================
#
# Architecture overview:
#
#   data/deliveries.json     — durable per-attempt records.
#   data/delivery-logs.jsonl — append-only audit trail.
#
#   Provider adapter         — providers.py. MockProvider today;
#                              Twilio/Meta WhatsApp Cloud drop in next.
#
#   Worker thread            — single background daemon that drains the
#                              Queued queue. Started on Flask boot, idempotent
#                              start/stop endpoints.
#
# Mapping integrity:
#   * Every delivery record carries the full recipient snapshot (name,
#     phone, address) at enqueue time. Renames / re-imports do NOT mutate
#     historical records.
#   * The link between a generated output and a recipient is the filename
#     stem (sanitize(address)), which both generators produce identically.
# ---------------------------------------------------------------------------

# In-process state.
_DELIVERIES_LOCK    = threading.Lock()
_DELIVERY_LOG_LOCK  = threading.Lock()
_WORKER_STOP_EVT    = threading.Event()
_WORKER_THREAD      = None
_WORKER_LOCK        = threading.Lock()

# Provider — defaults to MockProvider, OR auto-bootstraps WhatsApp if the
# required env vars are all set. Operator can still hot-swap via the
# POST /deliveries/provider endpoint.
# ---------------------------------------------------------------------------
# Two-step engagement flow config — read from env + persisted override file.
# Wrapped in a helper so the worker, enqueue path, and template-config
# endpoint all see the same answer.
# ---------------------------------------------------------------------------
def _flow_config():
    """Return the resolved flow + prompt-template settings.

    Layered: env vars provide the default, data/whatsapp-template.json
    overrides them at runtime so the operator can switch flow via the
    Delivery page without restarting.
    """
    env_flow   = (os.environ.get("WHATSAPP_FLOW") or "direct").strip().lower()
    env_prompt = (os.environ.get("WHATSAPP_PROMPT_TEMPLATE") or "").strip()
    env_lang   = (os.environ.get("WHATSAPP_PROMPT_LANG")
                  or os.environ.get("WHATSAPP_TEMPLATE_LANG")
                  or "en").strip() or "en"
    env_body   = (os.environ.get("WHATSAPP_PROMPT_BODY_PARAMS") or "").strip()
    prompt_body_params = [p.strip() for p in env_body.split(",") if p.strip()] if env_body else []

    try:
        override_path = os.path.join(PROJECT_ROOT, "data", "whatsapp-template.json")
        if os.path.isfile(override_path):
            with open(override_path, "r", encoding="utf-8") as _f:
                ov = json.load(_f) or {}
            if "flow" in ov:
                env_flow = (ov.get("flow") or env_flow).strip().lower()
            if "prompt_template" in ov:
                env_prompt = (ov.get("prompt_template") or "").strip()
            if "prompt_lang" in ov:
                env_lang   = (ov.get("prompt_lang") or env_lang).strip() or env_lang
            if "prompt_body_params" in ov:
                raw = ov.get("prompt_body_params") or []
                if isinstance(raw, str):
                    prompt_body_params = [s.strip() for s in raw.split(",") if s.strip()]
                else:
                    prompt_body_params = [str(x).strip() for x in raw if str(x).strip()]
    except Exception:  # noqa: BLE001
        pass

    if env_flow not in ("direct", "two-step"):
        env_flow = "direct"
    return {
        "flow":               env_flow,
        "prompt_template":    env_prompt,
        "prompt_lang":        env_lang,
        "prompt_body_params": prompt_body_params,
    }


def _bootstrap_provider():
    pid    = (os.environ.get("WHATSAPP_PHONE_NUMBER_ID") or "").strip()
    token  = (os.environ.get("WHATSAPP_ACCESS_TOKEN")    or "").strip()
    public = (os.environ.get("PUBLIC_BASE_URL")          or "").strip()
    if not public and os.environ.get("VERCEL") == "1":
        ver_url = os.environ.get("VERCEL_PROJECT_PRODUCTION_URL") or os.environ.get("VERCEL_URL")
        if ver_url:
            public = f"https://{ver_url}/api"
            os.environ["PUBLIC_BASE_URL"] = public
    if pid and token and public:
        try:
            kind = (os.environ.get("WHATSAPP_MEDIA_KIND") or "video").strip().lower()
            # Template config — when set, the provider sends a Marketing
            # template message instead of a freeform media message. This
            # is REQUIRED for business-initiated bulk sends; without it
            # Meta rejects with 131047 / 131049.
            tpl_image = (os.environ.get("WHATSAPP_TEMPLATE_IMAGE") or "").strip()
            tpl_video = (os.environ.get("WHATSAPP_TEMPLATE_VIDEO") or "").strip()
            tpl_lang  = (os.environ.get("WHATSAPP_TEMPLATE_LANG")  or "en").strip()
            # Body params: comma-separated format strings, each can
            # reference {name}/{address}/{phone}. Empty = no body params.
            body_raw  = (os.environ.get("WHATSAPP_TEMPLATE_BODY_PARAMS") or "").strip()
            body_params = [p.strip() for p in body_raw.split(",") if p.strip()] if body_raw else []
            # Persisted operator overrides (set via /deliveries/template-config)
            # take precedence over .env so UI changes survive restarts.
            try:
                override_path = os.path.join(PROJECT_ROOT, "data", "whatsapp-template.json")
                if os.path.isfile(override_path):
                    with open(override_path, "r", encoding="utf-8") as _f:
                        ov = json.load(_f) or {}
                    if "template_image" in ov:        tpl_image   = (ov.get("template_image") or "").strip()
                    if "template_video" in ov:        tpl_video   = (ov.get("template_video") or "").strip()
                    if "template_lang"  in ov:        tpl_lang    = (ov.get("template_lang")  or "en").strip() or "en"
                    if "template_body_params" in ov:
                        raw = ov.get("template_body_params") or []
                        if isinstance(raw, str):
                            body_params = [s.strip() for s in raw.split(",") if s.strip()]
                        else:
                            body_params = [str(x).strip() for x in raw if str(x).strip()]
            except Exception as _e:  # noqa: BLE001
                print(f"[startup] template override file read failed: {_e}", flush=True)
            print(
                f"[startup] WhatsApp templates: image={tpl_image or '(none)'} "
                f"video={tpl_video or '(none)'} lang={tpl_lang} "
                f"body_params={body_params}",
                flush=True,
            )
            if not tpl_image and not tpl_video:
                print(
                    "[startup] [WARN] No WhatsApp template configured - sends will "
                    "use freeform media which Meta REJECTS outside the 24h "
                    "customer-service window. Set WHATSAPP_TEMPLATE_IMAGE "
                    "and/or WHATSAPP_TEMPLATE_VIDEO in api/.env.",
                    flush=True,
                )
            return "whatsapp", get_provider(
                "whatsapp",
                phone_number_id=pid,
                access_token=token,
                public_base_url=public,
                media_kind=kind,
                template_image=tpl_image,
                template_video=tpl_video,
                template_lang=tpl_lang,
                template_body_params=body_params,
            )
        except Exception as e:  # noqa: BLE001
            print(f"[startup] WhatsApp provider bootstrap FAILED: {e} -- falling back to mock", flush=True)
    return "mock", get_provider("mock")

print(f"[cors] origins = {_cors_origins}", flush=True)
_PROVIDER_NAME, _PROVIDER = _bootstrap_provider()
# Wire the two-step flow's prompt-template attrs onto the live provider
# so send_prompt() works without a second constructor signature.
_flow = _flow_config()
if hasattr(_PROVIDER, "prompt_template"):
    _PROVIDER.prompt_template      = _flow["prompt_template"] or None
    _PROVIDER.prompt_lang          = _flow["prompt_lang"]
    _PROVIDER.prompt_body_params   = list(_flow["prompt_body_params"])
print(f"[startup] delivery provider: {_PROVIDER_NAME}", flush=True)
print(f"[startup] WhatsApp flow: {_flow['flow']}  prompt_template={_flow['prompt_template'] or '(none)'}  "
      f"prompt_lang={_flow['prompt_lang']}  prompt_body_params={_flow['prompt_body_params']}", flush=True)
if _flow["flow"] == "two-step" and not _flow["prompt_template"]:
    print("[startup] [WARN] flow=two-step but WHATSAPP_PROMPT_TEMPLATE is empty - "
          "stage-1 prompts will fail. Set the approved text template name in api/.env "
          "or via /deliveries/template-config.", flush=True)

# --------------------------------------------------------------------------
# Webhook & tunnel observability — printed once at boot so the operator
# can confirm the right config landed without digging through .env.
# --------------------------------------------------------------------------
def _startup_observability():
    pbu          = (os.environ.get("PUBLIC_BASE_URL") or "").strip()
    verify_token = (os.environ.get("WHATSAPP_WEBHOOK_VERIFY_TOKEN") or "").strip()
    app_secret   = (os.environ.get("WHATSAPP_APP_SECRET") or "").strip()
    webhook_url  = (pbu.rstrip("/") + "/deliveries/whatsapp-webhook") if pbu else "(PUBLIC_BASE_URL not set)"
    print("[startup] " + "=" * 60, flush=True)
    print(f"[startup] webhook_url           = {webhook_url}", flush=True)
    print(f"[startup] verify_token_present  = {bool(verify_token)}", flush=True)
    print(f"[startup] app_secret_present    = {bool(app_secret)}", flush=True)
    print(f"[startup] provider              = {_PROVIDER_NAME}", flush=True)
    # Persist last-seen PUBLIC_BASE_URL so the frontend can detect tunnel
    # rotation (TryCloudflare URLs change session-to-session) and prompt
    # the operator to update Meta's webhook config.
    #
    # CRITICAL: this write must be unconditional and unicode-safe — past
    # versions had a `⚠` glyph in the diagnostic print which blew up on
    # Windows cp1252 consoles, killed the function mid-flight, and left
    # the marker stale for days (CLAUDE.md tripwire #13). Marker write
    # therefore happens FIRST, then the diagnostic print, in its own
    # try-except so a print failure can't strand state.
    if pbu:
        try:
            marker_dir  = os.path.join(PROJECT_ROOT, "data")
            marker_path = os.path.join(marker_dir, "last-public-base-url.txt")
            os.makedirs(marker_dir, exist_ok=True)
            prev = ""
            if os.path.isfile(marker_path):
                with open(marker_path, "r", encoding="utf-8") as f:
                    prev = f.read().strip()
            with open(marker_path, "w", encoding="utf-8") as f:
                f.write(pbu)
            if prev and prev != pbu:
                try:
                    print(
                        f"[startup] [WARN] PUBLIC_BASE_URL CHANGED since last boot - "
                        f"update Meta webhook callback URL to: {webhook_url}",
                        flush=True,
                    )
                except Exception:
                    pass
        except Exception as e:  # noqa: BLE001
            print(f"[startup] rotation-detector warn: {e}", flush=True)
    print("[startup] " + "=" * 60, flush=True)

_startup_observability()

# Configuration knobs (could move to settings.json later).
_MAX_ATTEMPTS    = 3       # how many times to auto-retry inside the worker
_WORKER_IDLE_S   = 1.0     # sleep when queue is empty
_LOG_TAIL_BYTES  = 256_000 # safety cap when slurping the .jsonl tail


# ---------------------------------------------------------------------------
# Filename-stem helper — MUST match scripts/video_generator.py::sanitize_filename.
# ---------------------------------------------------------------------------
_SANITIZE_STRIP_RE = re.compile(r'[<>:"/\\|?*,]')
_SANITIZE_WS_RE    = re.compile(r"\s+")

def _sanitize_stem(text):
    """Mirror of the generators' sanitize_filename, used to link a
    recipient (by address) to the generated output files (by stem)."""
    s = _SANITIZE_WS_RE.sub("_", (text or "").strip())
    s = _SANITIZE_STRIP_RE.sub("", s)
    return s or "output1"


# ---------------------------------------------------------------------------
# Persistence helpers
# ---------------------------------------------------------------------------

def _empty_deliveries_doc():
    return {"version": 1, "updatedAt": 0, "items": []}


def _load_deliveries():
    if not os.path.isfile(DELIVERIES_PATH):
        return _empty_deliveries_doc()
    try:
        with open(DELIVERIES_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        return _empty_deliveries_doc()
    if not isinstance(data, dict):
        return _empty_deliveries_doc()
    items = data.get("items") or []
    if not isinstance(items, list):
        items = []
    return {
        "version":   int(data.get("version") or 1),
        "updatedAt": int(data.get("updatedAt") or 0),
        "items":     [it for it in items if isinstance(it, dict)],
    }


def _save_deliveries(doc):
    doc["updatedAt"] = int(time.time() * 1000)
    os.makedirs(os.path.dirname(DELIVERIES_PATH), exist_ok=True)
    tmp = DELIVERIES_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(doc, f, indent=2, ensure_ascii=False)
        f.write("\n")
    os.replace(tmp, DELIVERIES_PATH)


def _now_ms():
    return int(time.time() * 1000)


def _delivery_log(level, message, **fields):
    """Append a structured log line. Never raises into the caller."""
    line = {
        "ts":      _now_ms(),
        "level":   level,
        "message": message,
    }
    if fields:
        line.update(fields)
    try:
        with _DELIVERY_LOG_LOCK:
            os.makedirs(os.path.dirname(DELIVERY_LOG_PATH), exist_ok=True)
            with open(DELIVERY_LOG_PATH, "a", encoding="utf-8") as f:
                f.write(json.dumps(line, ensure_ascii=False) + "\n")
    except OSError:
        pass  # logging must never break the worker


def _read_delivery_log_tail(limit=100):
    """Return the last `limit` log entries, newest-last."""
    if not os.path.isfile(DELIVERY_LOG_PATH):
        return []
    try:
        size = os.path.getsize(DELIVERY_LOG_PATH)
        with open(DELIVERY_LOG_PATH, "rb") as f:
            if size > _LOG_TAIL_BYTES:
                f.seek(-_LOG_TAIL_BYTES, os.SEEK_END)
                # Discard the partial line we may have landed on.
                f.readline()
            raw = f.read()
    except OSError:
        return []
    out = []
    for line in raw.decode("utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return out[-limit:]


# ---------------------------------------------------------------------------
# Live media detection — the single source of truth for "what file should
# this recipient receive RIGHT NOW". Called from:
#   * the delivery view's cross-join (so the UI is always live)
#   * the worker's queue read (so the right file is shipped)
#   * the enqueue auto-mode (so new rows record the right kind)
#
# Trust order:
#   1. file actually on disk for the recipient stem
#   2. when both kinds exist, WHATSAPP_MEDIA_KIND env preference
#   3. neither exists → return None, caller decides what to do
#
# Never trusts the stored `media_kind` field on an existing delivery row
# — that's a snapshot, not the truth. This is what fixes the "row says
# Video but only the PNG exists" bug.
# ---------------------------------------------------------------------------

_IMAGE_OUT_EXTS_SET = _IMAGE_OUT_EXTS  # already a set; alias for clarity
_VIDEO_OUT_EXTS_SET = _VIDEO_OUT_EXTS


def _detect_media_for_stem(stem, *, videos=None, images=None, log=True):
    """Live filesystem detection. Returns a dict:
        { 'kind': 'image'|'video'|None,
          'filename': basename or None }

    Passes through `videos` / `images` (pre-scanned dicts) if the caller
    is iterating many stems, to avoid re-scanning the output dirs N times.
    """
    if videos is None:
        videos = _scan_output_dir(OUTPUT_VIDEOS, _VIDEO_OUT_EXTS_SET)
    if images is None:
        images = _scan_output_dir(OUTPUT_IMAGES, _IMAGE_OUT_EXTS_SET)
    v = videos.get(stem)
    i = images.get(stem)

    preferred = (os.environ.get("WHATSAPP_MEDIA_KIND") or "video").strip().lower()
    if preferred not in ("image", "video"):
        preferred = "video"

    if v and i:
        chosen = preferred
    elif v:
        chosen = "video"
    elif i:
        chosen = "image"
    else:
        chosen = None

    if chosen == "image":
        result = {"kind": "image", "filename": i["filename"]}
    elif chosen == "video":
        result = {"kind": "video", "filename": v["filename"]}
    else:
        result = {"kind": None, "filename": None}

    if log:
        print(
            f"[media-detect] recipient={stem} "
            f"image_found={bool(i)} video_found={bool(v)} "
            f"selected_type={result['kind']} selected_file={result['filename']}",
            flush=True,
        )
    return result


# ---------------------------------------------------------------------------
# Delivery construction
# ---------------------------------------------------------------------------

def _make_delivery(recipient, video_filename, image_filename, media_kind=None):
    """Build a fresh delivery record from a recipient + matching files.

    `media_kind` (optional) — `'image'` or `'video'` — tells the provider
    which file to ship to WhatsApp. None preserves legacy behavior
    (provider uses its constructor default).

    Two-step engagement flow: when WHATSAPP_FLOW=two-step (or override
    in data/whatsapp-template.json), the new delivery starts in
    `stage='prompt'` and the worker ships the text-only prompt template
    first. The webhook handler advances it to `stage='media'` when the
    recipient replies. Direct flow keeps `stage='media'` from the
    start so existing behavior is unchanged.
    """
    flow_cfg = _flow_config()
    is_two_step = (flow_cfg["flow"] == "two-step")
    return {
        "id":                  f"dlv_{uuid.uuid4().hex[:12]}",
        "stem":                _sanitize_stem(recipient.get("address", "")),
        "recipient_id":        recipient.get("id"),
        "recipient_name":      recipient.get("name", ""),
        "recipient_phone":     recipient.get("phone", ""),
        "recipient_address":   recipient.get("address", ""),
        "video_filename":      video_filename,
        "image_filename":      image_filename,
        # Per-delivery media kind. WhatsAppCloudProvider reads this first
        # and falls back to its constructor default if absent.
        "media_kind":          media_kind,
        # Two-step state machine columns. `flow` is snapshotted from
        # config AT ENQUEUE TIME so changing the config later doesn't
        # half-promote in-flight rows.
        "flow":                "two-step" if is_two_step else "direct",
        "stage":               "prompt"   if is_two_step else "media",
        "prompt_wamid":        None,
        "prompt_status":       None,
        "replied_at":          None,
        "inbound_message_id":  None,
        "status":              "Queued",
        "provider":            _PROVIDER.name,
        "provider_message_id": None,
        "attempts":            0,
        "max_attempts":        _MAX_ATTEMPTS,
        "last_error":          None,
        "createdAt":           _now_ms(),
        "updatedAt":           _now_ms(),
        "sentAt":              None,
        "deliveredAt":         None,
    }


def _resolve_auto_kind(recipient_stem, videos, images):
    """Thin wrapper around the centralized live detector — used by the
    enqueue path. Returns 'image' / 'video' / None."""
    det = _detect_media_for_stem(
        recipient_stem, videos=videos, images=images, log=False,
    )
    return det["kind"]


def _enqueue_recipients(recipient_subset, media_kind=None):
    """For each recipient, look up matching generated files and create a
    new Queued delivery. Returns a summary dict.

    `media_kind`:
        - 'image' / 'video' — explicit, only that kind enqueued; requires
          that file half on disk
        - None — AUTO mode: pick per-recipient based on
          WHATSAPP_MEDIA_KIND env preference, falling back to whatever
          kind is available. One delivery per recipient.

    Dedup is keyed by (recipient_id, stem, RESOLVED kind), so the same
    recipient can have an image AND a video delivery in flight without
    colliding, and auto-mode doesn't double-enqueue an existing row.
    """
    videos = _scan_output_dir(OUTPUT_VIDEOS, _VIDEO_OUT_EXTS)
    images = _scan_output_dir(OUTPUT_IMAGES, _IMAGE_OUT_EXTS)

    enqueued, skipped_missing, skipped_existing = [], [], []

    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        # Index by (recipient_id, stem, kind). Legacy rows without
        # media_kind end up keyed with None — they won't collide with
        # new explicit-kind rows.
        existing = {
            (d.get("recipient_id"), d.get("stem"), d.get("media_kind")): d
            for d in doc["items"]
        }

        for r in recipient_subset:
            stem = _sanitize_stem(r.get("address", ""))
            v = videos.get(stem)
            i = images.get(stem)

            # Resolve the effective kind. `media_kind=None` triggers
            # auto-mode (env preference + fallback).
            if media_kind in ("image", "video"):
                effective_kind = media_kind
                required = v if media_kind == "video" else i
                if not required:
                    skipped_missing.append({"recipient_id": r.get("id"),
                                            "stem": stem, "kind": media_kind})
                    continue
            else:
                effective_kind = _resolve_auto_kind(stem, videos, images)
                if not effective_kind:
                    skipped_missing.append({
                        "recipient_id": r.get("id"),
                        "stem":         stem,
                        "kind":         "auto",
                        "reason":       "neither image nor video on disk",
                    })
                    continue

            key = (r.get("id"), stem, effective_kind)
            if key in existing:
                skipped_existing.append({
                    "recipient_id": r.get("id"),
                    "stem":         stem,
                    "kind":         effective_kind,
                    "status":       existing[key].get("status"),
                })
                continue

            new = _make_delivery(
                r,
                v["filename"] if v else None,
                i["filename"] if i else None,
                media_kind=effective_kind,
            )
            doc["items"].append(new)
            enqueued.append(new)
            _delivery_log(
                "INFO", f"enqueued {effective_kind}",
                delivery_id=new["id"], stem=stem, kind=effective_kind,
                phone=new["recipient_phone"], name=new["recipient_name"],
            )

        if enqueued:
            _save_deliveries(doc)

    return {
        "enqueued":         len(enqueued),
        "items":            enqueued,
        "skipped_missing":  skipped_missing,
        "skipped_existing": skipped_existing,
    }


# ---------------------------------------------------------------------------
# Worker thread — sequential drain of the Queued queue.
# ---------------------------------------------------------------------------

def _take_next_queued():
    """Atomically pick the next Queued delivery and flip it to Sending.

    CRITICAL: re-runs `_detect_media_for_stem` against the LIVE filesystem
    here so a row stored months ago with `media_kind='video'` still ships
    the right file if the operator has since regenerated only the image.
    The row's `media_kind` / `*_filename` fields get rewritten in place
    so future reads see the corrected truth. This is the fix for "row
    says Video but only the PNG exists" → auto-corrects to image."""
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        target = None
        for d in doc["items"]:
            if d.get("status") == "Queued":
                target = d
                break
        if not target:
            return None

        # Live detection — auto-correct the stored kind/filename if the
        # filesystem has moved since enqueue.
        stem = target.get("stem", "")
        det  = _detect_media_for_stem(stem)
        if det["kind"]:
            # Rewrite stored fields to match reality. The provider will
            # read these freshly-corrected values.
            if target.get("media_kind") != det["kind"]:
                _delivery_log(
                    "INFO", "auto-correct kind",
                    delivery_id=target.get("id"), stem=stem,
                    was=target.get("media_kind"), now=det["kind"],
                )
                target["media_kind"] = det["kind"]
            if det["kind"] == "image":
                target["image_filename"] = det["filename"]
            else:
                target["video_filename"] = det["filename"]

        target["status"]    = "Sending"
        target["attempts"]  = int(target.get("attempts") or 0) + 1
        target["updatedAt"] = _now_ms()
        target["sentAt"]    = _now_ms()
        _save_deliveries(doc)
        # Return a shallow copy so the worker can use it without lock.
        return dict(target)


def _record_send_result(delivery_id, ok, provider_message_id, error, status=None):
    """Write the provider's verdict back into the delivery record.

    `status` (optional, async-provider escape hatch): pass `"Sending"` for
    providers like WhatsApp Cloud where send() success means "the API
    accepted the POST" but the real delivery outcome arrives later via a
    webhook. Defaults to `"Delivered"` on success — preserves the legacy
    sync-provider semantics (e.g. MockProvider) so nothing existing breaks.
    """
    if status not in (None, "Delivered", "Sending", "Pending Callback", "Media Sent"):
        status = None  # ignore unrecognised values rather than crash
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        for d in doc["items"]:
            if d.get("id") != delivery_id:
                continue
            attempts = int(d.get("attempts") or 0)
            d["updatedAt"] = _now_ms()
            d["provider"]  = _PROVIDER.name
            if ok:
                final_status = status or "Delivered"
                d["status"]              = final_status
                d["provider_message_id"] = provider_message_id
                d["last_error"]          = None
                if final_status == "Delivered":
                    d["deliveredAt"]     = _now_ms()
                else:
                    # Sending — webhook will set deliveredAt when Meta
                    # confirms; until then record when the POST went out.
                    d["sentAt"]          = _now_ms()
                _delivery_log(
                    "INFO", final_status.lower(),
                    delivery_id=d["id"], stem=d.get("stem"),
                    phone=d.get("recipient_phone"),
                    provider_message_id=provider_message_id,
                    attempts=attempts,
                )
            else:
                d["last_error"] = error or "unknown error"
                # Auto-requeue while we still have attempts left; otherwise
                # park as Failed so the operator can retry manually.
                if attempts < int(d.get("max_attempts") or _MAX_ATTEMPTS):
                    d["status"] = "Queued"
                    _delivery_log(
                        "WARN", "auto-requeue",
                        delivery_id=d["id"], stem=d.get("stem"),
                        phone=d.get("recipient_phone"),
                        attempts=attempts, error=d["last_error"],
                    )
                else:
                    d["status"] = "Failed"
                    _delivery_log(
                        "ERROR", "failed",
                        delivery_id=d["id"], stem=d.get("stem"),
                        phone=d.get("recipient_phone"),
                        attempts=attempts, error=d["last_error"],
                    )
            break
        _save_deliveries(doc)


# Stuck-Sending watchdog. WhatsApp's webhook normally arrives within
# 1-10s of the Graph POST. If we go this long without a status update,
# something silent broke (Meta's media fetch hung, tunnel dropped mid-
# request, recipient outside dev allowlist with no error webhook, etc.)
# Flip the row to Failed so the operator sees it instead of an
# indefinitely-yellow pill.
try:
    _SEND_TIMEOUT_S = int(os.environ.get("WHATSAPP_SEND_TIMEOUT_SECONDS", "180"))
except (TypeError, ValueError):
    _SEND_TIMEOUT_S = 180
_WATCHDOG_TICK_S = 15  # how often the worker checks for stuck rows


def _sweep_stuck_sending():
    """Rewritten watchdog. Silence from Meta's webhook is NOT a failure —
    it just means we're still waiting. So this sweep converts stale
    `Sending` rows (no callback yet) to `Pending Callback`, a non-terminal
    state that reads as "Meta accepted, awaiting confirmation".

    `Failed` is now reserved for EXPLICIT failure signals:
      * provider.send() returned ok=False (Graph error / network error)
      * Meta sent a `failed` status webhook
      * media URL fetch failed at Meta (also arrives as a `failed` webhook
        with a media-upload-error reason)

    Idempotent. Touches only rows older than `_SEND_TIMEOUT_S` with a
    wamid (= Meta accepted us). Rows in Sending without a wamid are
    pre-send hangs and stay Sending so the operator can spot them.
    """
    cutoff_ms = _now_ms() - _SEND_TIMEOUT_S * 1000
    flipped = []
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        changed = False
        for d in doc["items"]:
            if d.get("status") != "Sending":
                continue
            if not d.get("provider_message_id"):
                # No wamid means the provider hasn't successfully POSTed
                # yet — leave it as Sending so the operator notices a
                # genuinely stuck pre-send case.
                continue
            ts = d.get("sentAt") or d.get("updatedAt") or d.get("createdAt") or 0
            if ts > cutoff_ms:
                continue
            d["status"]    = "Pending Callback"
            d["updatedAt"] = _now_ms()
            flipped.append(d.get("id"))
            changed = True
        if changed:
            _save_deliveries(doc)
    for did in flipped:
        _delivery_log(
            "INFO", "watchdog: Sending -> Pending Callback "
                    "(no webhook yet, NOT a failure)",
            delivery_id=did,
        )


def _worker_loop():
    _delivery_log("INFO", "worker started", provider=_PROVIDER.name)
    last_sweep = 0
    while not _WORKER_STOP_EVT.is_set():
        # Stuck-Sending sweep on every tick (cheap — single in-memory scan).
        now = time.time()
        if now - last_sweep >= _WATCHDOG_TICK_S:
            try:
                _sweep_stuck_sending()
            except Exception as e:  # noqa: BLE001
                _delivery_log("ERROR", "watchdog sweep failed", error=str(e))
            last_sweep = now

        try:
            d = _take_next_queued()
        except Exception as e:
            _delivery_log("ERROR", "queue read failed", error=str(e))
            d = None
        if not d:
            # Nothing to do — wait but stay responsive to stop signals.
            _WORKER_STOP_EVT.wait(_WORKER_IDLE_S)
            continue

        # Branch on flow + stage. Two-step's stage-1 ships the prompt
        # template; stage-2 ships the actual media as freeform (allowed
        # inside the 24h customer-service window opened by the reply).
        flow_v  = (d.get("flow")  or "direct").lower()
        stage_v = (d.get("stage") or "media").lower()

        if flow_v == "two-step" and stage_v == "prompt":
            print(
                f"[prompt-stage] delivery_id={d.get('id')} phone={d.get('recipient_phone')} "
                f"prompt_template={getattr(_PROVIDER, 'prompt_template', None)!r}",
                flush=True,
            )
            try:
                result = _PROVIDER.send_prompt(d) if hasattr(_PROVIDER, "send_prompt") else {
                    "ok": False, "provider_message_id": None,
                    "error": "provider does not support two-step prompt sends",
                }
            except Exception as e:
                result = {"ok": False, "provider_message_id": None, "error": str(e)}
            _record_two_step_event(
                "last_prompt_send",
                delivery_id=d.get("id"),
                phone=d.get("recipient_phone"),
                prompt_wamid=result.get("provider_message_id"),
                ok=bool(result.get("ok")),
                error=result.get("error"),
            )
            _record_prompt_send_result(
                d["id"],
                ok=bool(result.get("ok")),
                prompt_wamid=result.get("provider_message_id"),
                error=result.get("error"),
            )
            continue

        # Direct flow (default) OR two-step stage-2 (media after reply).
        force_freeform = (flow_v == "two-step" and stage_v == "media")
        if force_freeform:
            print(
                f"[media-stage] delivery_id={d.get('id')} phone={d.get('recipient_phone')} "
                f"replied_at={d.get('replied_at')} stage=media force_freeform=True",
                flush=True,
            )
        try:
            if force_freeform and hasattr(_PROVIDER, "send"):
                # send() accepts the optional force_freeform kwarg on the
                # WhatsApp provider; mock + others ignore unknown kwargs
                # via their own signatures (Mock takes only delivery).
                try:
                    result = _PROVIDER.send(d, force_freeform=True)
                except TypeError:
                    result = _PROVIDER.send(d)
            else:
                result = _PROVIDER.send(d)
        except Exception as e:
            result = {"ok": False, "provider_message_id": None, "error": str(e)}
        if force_freeform:
            print(
                f"[media-send] delivery_id={d.get('id')} ok={result.get('ok')} "
                f"wamid={result.get('provider_message_id')} mode={result.get('send_mode')!r} "
                f"error={result.get('error')!r}",
                flush=True,
            )
            _record_two_step_event(
                "last_media_send",
                delivery_id=d.get("id"),
                phone=d.get("recipient_phone"),
                media_wamid=result.get("provider_message_id"),
                ok=bool(result.get("ok")),
                error=result.get("error"),
                mode=result.get("send_mode"),
            )
        _record_send_result(
            d["id"],
            ok=bool(result.get("ok")),
            provider_message_id=result.get("provider_message_id"),
            error=result.get("error"),
            # Async providers (WhatsApp) opt into Sending / Media Sent
            # here; sync providers (Mock) omit this and fall through to
            # Delivered.
            status=result.get("status"),
        )
    _delivery_log("INFO", "worker stopped")


def _record_prompt_send_result(delivery_id, ok, prompt_wamid, error):
    """Record the outcome of a stage-1 prompt-template send.

    Two-step state machine:
      * ok=True  → status = 'Awaiting Reply', prompt_wamid stored.
                   Row stays parked until the recipient replies — the
                   webhook handler's inbound parser advances it to
                   stage='media', status='Queued' for the worker to pick
                   up the actual media send.
      * ok=False → same auto-retry policy as the direct flow: requeue
                   the prompt while attempts remain, else mark Failed.
    """
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        for d in doc["items"]:
            if d.get("id") != delivery_id:
                continue
            attempts = int(d.get("attempts") or 0)
            d["updatedAt"] = _now_ms()
            d["provider"]  = _PROVIDER.name
            if ok:
                d["prompt_wamid"] = prompt_wamid
                d["status"]       = "Awaiting Reply"
                d["sentAt"]       = _now_ms()
                d["last_error"]   = None
                print(
                    f"[awaiting-reply] delivery_id={d.get('id')} phone={d.get('recipient_phone')} "
                    f"prompt_wamid={prompt_wamid}",
                    flush=True,
                )
                _delivery_log(
                    "INFO", "awaiting reply",
                    delivery_id=d["id"], phone=d.get("recipient_phone"),
                    prompt_wamid=prompt_wamid, attempts=attempts,
                )
            else:
                d["last_error"] = error or "unknown prompt-send error"
                if attempts < int(d.get("max_attempts") or _MAX_ATTEMPTS):
                    d["status"] = "Queued"  # try the prompt again
                    _delivery_log(
                        "WARN", "auto-requeue prompt",
                        delivery_id=d["id"], phone=d.get("recipient_phone"),
                        attempts=attempts, error=d["last_error"],
                    )
                else:
                    d["status"] = "Failed"
                    _delivery_log(
                        "ERROR", "prompt failed",
                        delivery_id=d["id"], phone=d.get("recipient_phone"),
                        attempts=attempts, error=d["last_error"],
                    )
            break
        _save_deliveries(doc)


def _advance_to_media_stage(phone_digits, inbound_wamid):
    """Match an inbound user message to an `Awaiting Reply` delivery
    and transition it to stage='media', status='Queued' so the worker
    ships the personalised media. Returns the number of rows advanced.

    Match is by `recipient_phone` normalised to digits-only. This is
    the same normalisation the provider applies before POST to Meta,
    so the round-trip is symmetric.
    """
    def _digits(s):
        return "".join(ch for ch in (s or "") if ch.isdigit())

    if not phone_digits:
        return 0
    touched = 0
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        for d in doc["items"]:
            if d.get("status") != "Awaiting Reply":
                continue
            if d.get("flow") != "two-step":
                continue
            if _digits(d.get("recipient_phone")) != phone_digits:
                continue
            # Transition: Awaiting Reply → Replied → Queued (stage=media).
            # We log the Replied state but immediately bump to Queued so
            # the worker picks it up on its next tick; the Replied label
            # only ever appears in the audit log line below.
            d["replied_at"]         = _now_ms()
            d["inbound_message_id"] = inbound_wamid
            d["stage"]              = "media"
            d["status"]             = "Queued"
            # Reset attempts so the media send gets its own retry budget.
            d["attempts"]           = 0
            d["last_error"]         = None
            d["updatedAt"]          = _now_ms()
            touched += 1
            print(
                f"[media-stage] delivery_id={d.get('id')} phone={d.get('recipient_phone')} "
                f"transition=awaiting_reply->queued_media inbound_wamid={inbound_wamid}",
                flush=True,
            )
            _record_two_step_event(
                "last_media_transition",
                delivery_id=d.get("id"),
                phone=d.get("recipient_phone"),
                inbound_wamid=inbound_wamid,
            )
            _delivery_log(
                "INFO", "replied -> queued for media",
                delivery_id=d.get("id"), phone=d.get("recipient_phone"),
                inbound_wamid=inbound_wamid,
            )
        if touched:
            _save_deliveries(doc)
    return touched


def _apply_prompt_status(prompt_wamid, meta_status, error_msg=None):
    """Status callback for a stage-1 prompt wamid. Updates the row's
    `prompt_status` field but does NOT change the main `status` (which
    is being driven by the two-step state machine). Returns True iff a
    matching row was found."""
    if not prompt_wamid:
        return False
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        touched = False
        for d in doc["items"]:
            if d.get("prompt_wamid") != prompt_wamid:
                continue
            d["prompt_status"] = meta_status
            d["updatedAt"]     = _now_ms()
            if meta_status == "failed":
                d["last_error"] = error_msg or "prompt send failed (per webhook)"
                # Promote to Failed only if we never advanced past prompt.
                # If the recipient already replied and we're in media
                # stage, ignore — Meta sometimes retries old status events.
                if d.get("stage") == "prompt":
                    d["status"] = "Failed"
            touched = True
            _delivery_log(
                "INFO" if meta_status != "failed" else "ERROR",
                f"prompt webhook {meta_status}",
                delivery_id=d.get("id"),
                prompt_wamid=prompt_wamid,
                error=error_msg,
            )
            break
        if touched:
            _save_deliveries(doc)
    return touched


def _start_worker():
    """Idempotent: returns whether a new thread was started."""
    global _WORKER_THREAD
    with _WORKER_LOCK:
        if _WORKER_THREAD and _WORKER_THREAD.is_alive():
            return False
        _WORKER_STOP_EVT.clear()
        _WORKER_THREAD = threading.Thread(target=_worker_loop, daemon=True, name="delivery-worker")
        _WORKER_THREAD.start()
        return True


def _stop_worker():
    """Signal stop. Thread exits on its next idle tick."""
    _WORKER_STOP_EVT.set()


def _worker_status():
    with _WORKER_LOCK:
        alive = bool(_WORKER_THREAD and _WORKER_THREAD.is_alive())
    return {
        "alive":     alive,
        "stop_set":  _WORKER_STOP_EVT.is_set(),
        "provider":  _PROVIDER.name,
        "providers": list_providers(),
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/deliveries", methods=["GET"])
def deliveries_list():
    """Full deliveries table + aggregate counts."""
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
    items = sorted(doc["items"], key=lambda d: d.get("updatedAt", 0), reverse=True)
    counts = {
        "Queued": 0, "Sending": 0, "Delivered": 0, "Failed": 0,
        # Two-step engagement flow states.
        "Awaiting Reply": 0, "Media Sent": 0,
    }
    for d in items:
        s = d.get("status")
        if s in counts:
            counts[s] += 1
    return jsonify({
        "status":    "success",
        "count":     len(items),
        "counts":    counts,
        "items":     items,
        "updatedAt": doc.get("updatedAt", 0),
        "worker":    _worker_status(),
    })


@app.route("/deliveries/enqueue", methods=["POST"])
def deliveries_enqueue():
    """Enqueue a subset of recipients (by id list). Body: {recipient_ids:[...]}"""
    payload = request.get_json(silent=True) or {}
    ids = payload.get("recipient_ids")
    if not isinstance(ids, list) or not ids:
        return jsonify({"status": "error", "error": "recipient_ids[] required"}), 400
    with _RECIPIENTS_LOCK:
        rdoc = _load_recipients()
    subset = [r for r in rdoc["items"] if r.get("id") in set(ids)]
    if not subset:
        return jsonify({"status": "error", "error": "no matching recipients found"}), 404
    result = _enqueue_recipients(subset)
    return jsonify({"status": "success", **result})


@app.route("/deliveries/enqueue-all", methods=["POST"])
def deliveries_enqueue_all():
    """Enqueue recipients for delivery.

    Body (all optional):
      { limit: N,            // Send first N (omit/0 = all)
        kind:  'image'|'video' }  // Which media to send; omit for legacy
                                  // paired behavior (requires both halves)
    """
    payload = request.get_json(silent=True) or {}
    limit = payload.get("limit")
    try:
        limit = int(limit) if limit not in (None, "", 0) else 0
    except (TypeError, ValueError):
        return jsonify({"status": "error", "error": "limit must be a positive integer"}), 400
    if limit < 0:
        return jsonify({"status": "error", "error": "limit must be non-negative"}), 400

    kind_raw = payload.get("kind")
    kind = None
    if kind_raw is not None:
        kind = str(kind_raw).strip().lower()
        if kind not in ("image", "video"):
            return jsonify({
                "status": "error",
                "error":  "kind must be 'image' or 'video' (omit for paired send)",
            }), 400

    with _RECIPIENTS_LOCK:
        rdoc = _load_recipients()
    pool = rdoc["items"]
    if limit:
        pool = pool[:limit]
    result = _enqueue_recipients(pool, media_kind=kind)
    return jsonify({"status": "success", "limit": limit, "kind": kind, **result})


@app.route("/deliveries/<dlv_id>/retry", methods=["POST"])
def deliveries_retry_one(dlv_id):
    """Reset a single delivery from Failed (or any state) back to Queued."""
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        target = None
        for d in doc["items"]:
            if d.get("id") == dlv_id:
                target = d
                break
        if not target:
            return jsonify({"status": "error", "error": "delivery not found"}), 404
        # Reset attempts so the worker can run the auto-retry chain again.
        target["status"]     = "Queued"
        target["attempts"]   = 0
        target["last_error"] = None
        target["sentAt"]     = None
        target["deliveredAt"] = None
        target["updatedAt"]  = _now_ms()
        _save_deliveries(doc)
    _delivery_log("INFO", "manual retry", delivery_id=dlv_id)
    return jsonify({"status": "success", "item": target})


@app.route("/deliveries/retry-failed", methods=["POST"])
def deliveries_retry_failed():
    """Bulk-requeue every Failed delivery."""
    requeued_ids = []
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        for d in doc["items"]:
            if d.get("status") == "Failed":
                d["status"]      = "Queued"
                d["attempts"]    = 0
                d["last_error"]  = None
                d["sentAt"]      = None
                d["deliveredAt"] = None
                d["updatedAt"]   = _now_ms()
                requeued_ids.append(d["id"])
        if requeued_ids:
            _save_deliveries(doc)
    for did in requeued_ids:
        _delivery_log("INFO", "bulk retry", delivery_id=did)
    return jsonify({"status": "success", "requeued": len(requeued_ids), "ids": requeued_ids})


@app.route("/deliveries/<dlv_id>", methods=["DELETE"])
def deliveries_delete(dlv_id):
    """Remove a single delivery record (audit cleanup)."""
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        before = len(doc["items"])
        doc["items"] = [d for d in doc["items"] if d.get("id") != dlv_id]
        if len(doc["items"]) == before:
            return jsonify({"status": "error", "error": "delivery not found"}), 404
        _save_deliveries(doc)
    return jsonify({"status": "success", "count": len(doc["items"])})


@app.route("/deliveries/clear", methods=["POST"])
def deliveries_clear():
    """Wipe the whole deliveries table — handy for resetting test runs."""
    payload = request.get_json(silent=True) or {}
    confirm = payload.get("confirm")
    if confirm != "yes":
        return jsonify({"status": "error", "error": "pass {confirm: 'yes'} to clear"}), 400
    with _DELIVERIES_LOCK:
        doc = _empty_deliveries_doc()
        _save_deliveries(doc)
    _delivery_log("INFO", "deliveries cleared")
    return jsonify({"status": "success"})


@app.route("/deliveries/worker", methods=["GET"])
def deliveries_worker_status():
    return jsonify({"status": "success", "worker": _worker_status()})


@app.route("/deliveries/worker/start", methods=["POST"])
def deliveries_worker_start():
    started = _start_worker()
    return jsonify({"status": "success", "started": started, "worker": _worker_status()})


@app.route("/deliveries/worker/stop", methods=["POST"])
def deliveries_worker_stop():
    _stop_worker()
    return jsonify({"status": "success", "worker": _worker_status()})


# ---------------------------------------------------------------------------
# WhatsApp template config — runtime override of which approved template
# the provider uses. Persisted to data/whatsapp-template.json so operator
# choices survive restart; that file overrides the WHATSAPP_TEMPLATE_*
# values in api/.env (env stays as the fallback when the file is absent).
# ---------------------------------------------------------------------------
def _template_config_path():
    return os.path.join(PROJECT_ROOT, "data", "whatsapp-template.json")


def _read_template_config_file():
    """Load the persisted template config, or {} if the file is absent."""
    path = _template_config_path()
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:  # noqa: BLE001
        return {}


def _write_template_config_file(cfg):
    path = _template_config_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)
    os.replace(tmp, path)


def _effective_template_config():
    """Merge env vars with the persisted JSON file (file wins). This is
    what the provider should see — the operator's UI override beats the
    .env baseline, but env fills in any keys the file doesn't set."""
    env = {
        "template_image":       (os.environ.get("WHATSAPP_TEMPLATE_IMAGE") or "").strip(),
        "template_video":       (os.environ.get("WHATSAPP_TEMPLATE_VIDEO") or "").strip(),
        "template_lang":        (os.environ.get("WHATSAPP_TEMPLATE_LANG")  or "en").strip(),
        "flow":                 (os.environ.get("WHATSAPP_FLOW") or "direct").strip().lower(),
        "prompt_template":      (os.environ.get("WHATSAPP_PROMPT_TEMPLATE") or "").strip(),
        "prompt_lang":          (os.environ.get("WHATSAPP_PROMPT_LANG")
                                 or os.environ.get("WHATSAPP_TEMPLATE_LANG")
                                 or "en").strip(),
    }
    body_raw = (os.environ.get("WHATSAPP_TEMPLATE_BODY_PARAMS") or "").strip()
    env["template_body_params"] = (
        [p.strip() for p in body_raw.split(",") if p.strip()] if body_raw else []
    )
    p_body_raw = (os.environ.get("WHATSAPP_PROMPT_BODY_PARAMS") or "").strip()
    env["prompt_body_params"] = (
        [p.strip() for p in p_body_raw.split(",") if p.strip()] if p_body_raw else []
    )
    file_cfg = _read_template_config_file()
    merged = {**env, **{k: v for k, v in file_cfg.items() if v is not None}}
    if merged.get("flow") not in ("direct", "two-step"):
        merged["flow"] = "direct"
    return merged


def _apply_template_config_to_provider(cfg):
    """Mutate the live _PROVIDER attributes so the next send picks up the
    new template config without a process restart. Mock provider has none
    of these attrs — guard with hasattr() so a hot-swapped mock survives."""
    global _PROVIDER
    p = _PROVIDER
    if hasattr(p, "template_image"):
        p.template_image       = (cfg.get("template_image") or "").strip() or None
    if hasattr(p, "template_video"):
        p.template_video       = (cfg.get("template_video") or "").strip() or None
    if hasattr(p, "template_lang"):
        p.template_lang        = (cfg.get("template_lang")  or "en").strip() or "en"
    if hasattr(p, "template_body_params"):
        params = cfg.get("template_body_params") or []
        if isinstance(params, str):
            params = [s.strip() for s in params.split(",") if s.strip()]
        p.template_body_params = list(params)
    # Two-step prompt-template config — live-update so the next stage-1
    # send uses the new values without restart.
    if hasattr(p, "prompt_template"):
        p.prompt_template = (cfg.get("prompt_template") or "").strip() or None
    if hasattr(p, "prompt_lang"):
        p.prompt_lang = (cfg.get("prompt_lang") or cfg.get("template_lang") or "en").strip() or "en"
    if hasattr(p, "prompt_body_params"):
        params = cfg.get("prompt_body_params") or []
        if isinstance(params, str):
            params = [s.strip() for s in params.split(",") if s.strip()]
        p.prompt_body_params = list(params)


@app.route("/deliveries/diagnostics", methods=["GET"])
def deliveries_diagnostics():
    """Live snapshot of the two-step engagement flow's state, for the
    Delivery page's diagnostics panel.

    Returns the active flow, whether the prompt template is configured,
    counts per state-machine state, and the most recent two-step events
    (last inbound message, last awaiting→media transition, last prompt
    send, last media send, last media delivered). All best-effort
    in-memory snapshots — safe to call from any frontend poll cadence.
    """
    eff = _effective_template_config()
    flow = eff.get("flow") or "direct"
    prompt_configured = bool((eff.get("prompt_template") or "").strip())
    with _TWO_STEP_DIAG_LOCK:
        diag = dict(_TWO_STEP_DIAG)

    # Per-state counts from the live deliveries store.
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        counts = {}
        for d in doc.get("items", []):
            s = d.get("status") or "Queued"
            counts[s] = counts.get(s, 0) + 1
            # Surface per-stage breakdown for two-step rows so the panel
            # can show "3 awaiting reply, 1 in media stage".
        # Count the two-step substates explicitly.
        awaiting_reply = sum(1 for d in doc.get("items", []) if d.get("status") == "Awaiting Reply")
        media_stage_q  = sum(1 for d in doc.get("items", []) if d.get("flow") == "two-step"
                             and d.get("stage") == "media" and d.get("status") in ("Queued", "Sending"))

    # `messages_subscription_detected` — we have no way to query Meta's
    # dashboard, but if we have EVER received an inbound message at the
    # current PUBLIC_BASE_URL the field must be subscribed. Returns true
    # only after the first inbound; until then we leave it unknown.
    last_inbound = diag.get("last_inbound")
    pbu = (os.environ.get("PUBLIC_BASE_URL") or "").strip()
    sync_url = _read_webhook_sync()
    messages_subscription_detected = bool(last_inbound) and bool(pbu) and (sync_url == pbu)

    return jsonify({
        "status":             "success",
        "flow":               flow,
        "prompt_template":    eff.get("prompt_template") or "",
        "prompt_configured":  prompt_configured,
        "callback_url":       (pbu.rstrip("/") + "/deliveries/whatsapp-webhook") if pbu else None,
        "messages_subscription_detected": messages_subscription_detected,
        "counts":             counts,
        "two_step_counts":    {
            "awaiting_reply":  awaiting_reply,
            "media_in_flight": media_stage_q,
        },
        "last_inbound":           diag.get("last_inbound"),
        "last_media_transition":  diag.get("last_media_transition"),
        "last_prompt_send":       diag.get("last_prompt_send"),
        "last_media_send":        diag.get("last_media_send"),
        "last_media_delivered":   diag.get("last_media_delivered"),
    })


@app.route("/deliveries/<delivery_id>/simulate-reply", methods=["POST"])
def deliveries_simulate_reply(delivery_id):
    """Debug-only: synthesise an inbound user reply for the given
    delivery and advance the two-step state machine.

    Same code path as a real Meta inbound webhook — it just bypasses
    Meta's signature check and constructs the inbound wamid locally.
    Useful for operator-driven end-to-end tests without needing a real
    recipient to respond. Returns the number of rows advanced (0 if the
    target row isn't in Awaiting Reply state or doesn't exist).
    """
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        target = next((d for d in doc["items"] if d.get("id") == delivery_id), None)
    if not target:
        return jsonify({"status": "error", "error": "delivery_id not found"}), 404
    if target.get("status") != "Awaiting Reply":
        return jsonify({
            "status": "error",
            "error":  f"delivery is in {target.get('status')!r}, not Awaiting Reply",
        }), 400

    phone = target.get("recipient_phone") or ""
    digits = "".join(ch for ch in phone if ch.isdigit())
    synth_wamid = f"wamid.SIM_{uuid.uuid4().hex[:16]}"
    print(
        f"[inbound] (SIMULATED) from={phone} digits={digits} type=text "
        f"id={synth_wamid} preview='YES (operator-simulated)'",
        flush=True,
    )
    _record_two_step_event(
        "last_inbound",
        phone=phone, digits=digits, msg_id=synth_wamid,
        msg_type="text", preview="YES (operator-simulated)",
        simulated=True,
    )
    advanced = _advance_to_media_stage(digits, synth_wamid)
    return jsonify({
        "status":           "success",
        "advanced":         advanced,
        "synthetic_wamid":  synth_wamid,
    })


@app.route("/deliveries/template-config", methods=["GET", "POST"])
def deliveries_template_config():
    """Read or update the active WhatsApp template configuration.

    GET  → returns the current effective config (env + persisted file).
    POST → body { template_image, template_video, template_lang,
                  template_body_params } updates persisted config AND the
                  live provider. Any field omitted from the POST body is
                  left unchanged. Body params can be sent as a list of
                  strings OR a single comma-separated string.

    The persisted config file is read by `_bootstrap_provider()` on
    startup so operator choices survive process restarts.
    """
    if request.method == "GET":
        eff = _effective_template_config()
        # `modes` describes the per-kind active SEND mode. In two-step
        # flow the *first* hop is always the prompt template; the media
        # hop happens only after a reply, so we surface both.
        is_two_step = eff.get("flow") == "two-step"
        if is_two_step:
            stage1 = f"prompt:{eff.get('prompt_template')}" if eff.get("prompt_template") else "prompt:(unset)"
            modes = {
                "image": f"{stage1} → freeform-window:image",
                "video": f"{stage1} → freeform-window:video",
            }
        else:
            modes = {
                "image": ("template:" + eff["template_image"]) if eff.get("template_image") else "freeform",
                "video": ("template:" + eff["template_video"]) if eff.get("template_video") else "freeform",
            }
        return jsonify({
            "status":   "success",
            "config":   eff,
            "modes":    modes,
            "provider": _PROVIDER_NAME,
        })

    payload = request.get_json(silent=True) or {}
    current = _read_template_config_file()
    # Whitelist accepted keys so the operator can't smuggle extras into
    # the persisted file.
    for key in ("template_image", "template_video", "template_lang",
                "prompt_template", "prompt_lang"):
        if key in payload:
            val = payload[key]
            current[key] = ("" if val is None else str(val)).strip()
    if "flow" in payload:
        f = (str(payload.get("flow") or "")).strip().lower()
        current["flow"] = f if f in ("direct", "two-step") else "direct"
    if "template_body_params" in payload:
        raw = payload["template_body_params"]
        if isinstance(raw, list):
            current["template_body_params"] = [str(x).strip() for x in raw if str(x).strip()]
        elif isinstance(raw, str):
            current["template_body_params"] = [s.strip() for s in raw.split(",") if s.strip()]
        else:
            current["template_body_params"] = []
    if "prompt_body_params" in payload:
        raw = payload["prompt_body_params"]
        if isinstance(raw, list):
            current["prompt_body_params"] = [str(x).strip() for x in raw if str(x).strip()]
        elif isinstance(raw, str):
            current["prompt_body_params"] = [s.strip() for s in raw.split(",") if s.strip()]
        else:
            current["prompt_body_params"] = []
    try:
        _write_template_config_file(current)
    except Exception as e:  # noqa: BLE001
        return jsonify({"status": "error", "error": f"could not persist: {e}"}), 500

    eff = _effective_template_config()
    _apply_template_config_to_provider(eff)
    _delivery_log("INFO", "template config updated",
                  image=eff.get("template_image") or "",
                  video=eff.get("template_video") or "")
    return jsonify({"status": "success", "config": eff})


@app.route("/deliveries/provider", methods=["GET", "POST"])
def deliveries_provider():
    """GET → current + available providers. POST {name, **kwargs} → switch
    provider. New deliveries (and any in-flight ones) use the new provider
    on their next send attempt."""
    global _PROVIDER, _PROVIDER_NAME
    if request.method == "GET":
        return jsonify({
            "status":  "success",
            "active":  _PROVIDER_NAME,
            "available": list_providers(),
        })
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip().lower()
    if not name:
        return jsonify({"status": "error", "error": "name required"}), 400
    try:
        kwargs = {k: v for k, v in payload.items() if k != "name"}
        new_provider = get_provider(name, **kwargs)
    except (ValueError, TypeError) as e:
        return jsonify({"status": "error", "error": str(e)}), 400
    _PROVIDER      = new_provider
    _PROVIDER_NAME = name
    _delivery_log("INFO", "provider switched", provider=name)
    return jsonify({"status": "success", "active": name})


# ---------------------------------------------------------------------------
# WhatsApp Business Cloud API — webhook endpoints
#
# Meta calls these unauthenticated (from their CDN, from arbitrary IPs).
# Both routes are in _PUBLIC_PREFIXES so the auth middleware doesn't reject
# them. The POST handler enforces cryptographic verification of every
# incoming hit via the X-Hub-Signature-256 header
# (HMAC-SHA256 of the raw body, key = WHATSAPP_APP_SECRET).
#
# GET  → verify-token challenge (Meta sends this once when you add the
#        webhook URL to your app). We echo `hub.challenge` iff
#        `hub.verify_token` matches WHATSAPP_WEBHOOK_VERIFY_TOKEN.
# POST → status callbacks. Body shape per
#        https://developers.facebook.com/docs/whatsapp/cloud-api/webhooks/payload-examples
# ---------------------------------------------------------------------------
def _verify_meta_signature(raw_body):
    """Constant-time HMAC-SHA256 check. Returns True iff the request's
    X-Hub-Signature-256 header matches `sha256=<hex(HMAC(secret, body))>`.
    If WHATSAPP_APP_SECRET isn't configured we REJECT (fail-closed)."""
    secret = (os.environ.get("WHATSAPP_APP_SECRET") or "").strip()
    if not secret:
        return False
    header = request.headers.get("X-Hub-Signature-256", "")
    if not header.startswith("sha256="):
        return False
    received = header.split("=", 1)[1].strip()
    expected = hmac.new(
        secret.encode("utf-8"),
        raw_body,
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected, received)


# Meta WhatsApp Cloud → our internal delivery row status.
#   sent       = Meta edge accepted, in transit  → Pending Callback
#                (downgrade from Delivered NOT allowed; protects ordering)
#   delivered  = recipient device received
#   read       = recipient opened (tracked as its own state for diagnostics)
#   failed     = permanent failure (worker won't auto-retry)
_WHATSAPP_STATUS_MAP = {
    "sent":      "Pending Callback",
    "delivered": "Delivered",
    "read":      "Read",
    "failed":    "Failed",
}


# In-memory tracker of the most recent webhook hit. Surfaced by the
# /deliveries/webhook-status endpoint so the Delivery page can show
# "Last callback X seconds ago from IP Y, status Z".
_LAST_WEBHOOK = {
    "received_at_ms": None,   # epoch ms when we received the POST
    "ip":             None,   # request.remote_addr (after ProxyFix)
    "wamid":          None,   # most recent provider_message_id we processed
    "meta_status":    None,   # 'sent' / 'delivered' / 'read' / 'failed'
    "error":          None,
    "verify_at_ms":   None,   # last successful GET verify-challenge
}
_LAST_WEBHOOK_LOCK = threading.Lock()

# In-memory trackers for the Delivery page diagnostics panel. Records
# the most recent two-step state-machine event of each type. Read by
# /deliveries/diagnostics; safe to lose on restart (purely informational).
_TWO_STEP_DIAG = {
    "last_inbound":         None,  # { at_ms, phone, msg_id, type, preview }
    "last_media_transition": None, # { at_ms, delivery_id, phone, inbound_wamid }
    "last_prompt_send":     None,  # { at_ms, delivery_id, phone, prompt_wamid }
    "last_media_send":      None,  # { at_ms, delivery_id, phone, media_wamid }
    "last_media_delivered": None,  # { at_ms, delivery_id, phone, media_wamid }
}
_TWO_STEP_DIAG_LOCK = threading.Lock()


def _record_two_step_event(slot, **payload):
    """Best-effort update of one of the _TWO_STEP_DIAG slots. Never raises
    into the caller; pure observability."""
    try:
        with _TWO_STEP_DIAG_LOCK:
            _TWO_STEP_DIAG[slot] = {"at_ms": _now_ms(), **payload}
    except Exception:  # noqa: BLE001
        pass


def _apply_webhook_status(wamid, meta_status, timestamp_s, error_msg=None):
    """Find the delivery row with this provider_message_id and apply the
    Meta status update. Returns True iff a row was touched."""
    target_status = _WHATSAPP_STATUS_MAP.get(meta_status)
    if not target_status:
        return False  # unrecognised status — ignore
    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
        touched = False
        for d in doc["items"]:
            if d.get("provider_message_id") != wamid:
                continue
            current = d.get("status")
            # State machine — refuse downgrades. Meta sometimes sends
            # events out of order (e.g. `sent` arriving AFTER `delivered`).
            # Progression weight: terminal Failed > Read > Delivered >
            # Pending Callback > Sending > Queued. We never demote.
            order = {
                "Queued": 0, "Awaiting Reply": 0, "Sending": 1,
                "Pending Callback": 2, "Media Sent": 2,
                "Delivered": 3, "Read": 4, "Failed": 5,
            }
            if order.get(target_status, 0) < order.get(current, 0):
                touched = True  # log but don't downgrade
                _delivery_log(
                    "INFO", f"webhook {meta_status} ignored (no downgrade)",
                    delivery_id=d.get("id"), current=current,
                    proposed=target_status, provider_message_id=wamid,
                )
                break
            d["status"]    = target_status
            d["updatedAt"] = _now_ms()
            if target_status == "Delivered" and not d.get("deliveredAt"):
                d["deliveredAt"] = int((timestamp_s or time.time()) * 1000)
                # Tag two-step deliveries so the diagnostics panel can show
                # the closing-the-loop event. Direct flow logs are unaffected.
                if d.get("flow") == "two-step" and d.get("provider_message_id") == wamid:
                    print(
                        f"[media-delivered] delivery_id={d.get('id')} phone={d.get('recipient_phone')} "
                        f"wamid={wamid}",
                        flush=True,
                    )
                    _record_two_step_event(
                        "last_media_delivered",
                        delivery_id=d.get("id"),
                        phone=d.get("recipient_phone"),
                        media_wamid=wamid,
                    )
            if target_status == "Failed":
                d["last_error"] = error_msg or "delivery failed (per WhatsApp webhook)"
            touched = True
            _delivery_log(
                "INFO" if target_status != "Failed" else "ERROR",
                f"webhook {meta_status}",
                delivery_id=d.get("id"),
                stem=d.get("stem"),
                phone=d.get("recipient_phone"),
                provider_message_id=wamid,
                error=error_msg,
            )
            break
        if touched:
            _save_deliveries(doc)
    return touched


# ---------------------------------------------------------------------------
# Webhook-sync marker.
#
# Records the PUBLIC_BASE_URL that was active the LAST time Meta successfully
# reached our webhook (either the verify GET or a status POST). This is what
# the frontend "Webhook URL changed" banner SHOULD compare against — not the
# startup-time marker — because we want the banner to disappear the moment
# Meta has been reconfigured to call the new URL, even before any restart.
#
# File: data/webhook-sync-at-url.txt (single line, the URL or empty).
# ---------------------------------------------------------------------------
def _webhook_sync_marker_path():
    return os.path.join(PROJECT_ROOT, "data", "webhook-sync-at-url.txt")


def _record_webhook_sync():
    """Mark that Meta just reached us at the current PUBLIC_BASE_URL.
    Best-effort — never raises into the request handler."""
    pbu = (os.environ.get("PUBLIC_BASE_URL") or "").strip()
    if not pbu:
        return
    try:
        path = _webhook_sync_marker_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(pbu)
    except Exception:  # noqa: BLE001
        pass


def _read_webhook_sync():
    """Return the URL recorded at last successful Meta callback, or None."""
    try:
        path = _webhook_sync_marker_path()
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                return f.read().strip() or None
    except Exception:  # noqa: BLE001
        pass
    return None


def _header_summary():
    """Compact, secret-safe dump of request headers — for log breadcrumbs."""
    interesting = (
        "User-Agent", "Content-Type", "Content-Length",
        "X-Forwarded-For", "X-Forwarded-Proto", "X-Forwarded-Host",
        "X-Hub-Signature-256",
    )
    out = {}
    for k in interesting:
        v = request.headers.get(k)
        if not v:
            continue
        # Don't echo the full signature; just length + prefix.
        if k == "X-Hub-Signature-256":
            v = f"sha256=<{len(v)} chars, prefix={v[7:15] if len(v) > 15 else v}>"
        out[k] = v
    return out


@app.route("/deliveries/whatsapp-webhook", methods=["GET"])
def deliveries_whatsapp_webhook_verify():
    """Meta's one-time webhook verification handshake.
    See: https://developers.facebook.com/docs/graph-api/webhooks/getting-started
    """
    expected = (os.environ.get("WHATSAPP_WEBHOOK_VERIFY_TOKEN") or "").strip()
    mode      = request.args.get("hub.mode")
    token     = request.args.get("hub.verify_token") or ""
    challenge = request.args.get("hub.challenge") or ""

    print(
        f"[webhook-verify] method=GET ip={_client_ip()} mode={mode} "
        f"token_match={bool(expected) and hmac.compare_digest(token, expected)} "
        f"headers={_header_summary()}",
        flush=True,
    )

    if mode == "subscribe" and expected and hmac.compare_digest(token, expected):
        with _LAST_WEBHOOK_LOCK:
            _LAST_WEBHOOK["verify_at_ms"] = _now_ms()
            _LAST_WEBHOOK["ip"]           = _client_ip()
        # Meta just spoke to us at the current PUBLIC_BASE_URL — record it
        # so the frontend banner clears immediately on the next poll.
        _record_webhook_sync()
        print(f"[webhook-verify] OK - echoing challenge ({len(challenge)} chars)", flush=True)
        return challenge, 200, {"Content-Type": "text/plain"}
    print(f"[webhook-verify] REJECTED (mode={mode}, expected_token_set={bool(expected)})", flush=True)
    return jsonify({"status": "error", "error": "forbidden"}), 403


@app.route("/deliveries/whatsapp-webhook", methods=["POST"])
def deliveries_whatsapp_webhook():
    """Status callback receiver. Meta retries on non-2xx so we MUST return
    200 once we've processed (or knowingly ignored) the payload — even if
    individual entries inside the batch are malformed."""
    raw = request.get_data() or b""
    ip  = _client_ip()

    # Loud-but-safe log of EVERY incoming POST. The signature check below
    # gates side effects; we log every attempt so a missing webhook
    # (causes #1/#3 from the diagnosis) is visible immediately.
    print(
        f"[webhook-callback] method=POST ip={ip} body_bytes={len(raw)} "
        f"headers={_header_summary()}",
        flush=True,
    )

    if not _verify_meta_signature(raw):
        print(f"[webhook-callback] signature MISMATCH from {ip}", flush=True)
        return jsonify({"status": "error", "error": "invalid signature"}), 403

    # Signature verified — Meta is talking to the current PUBLIC_BASE_URL.
    # Record this so the rotation banner clears even if no status entries
    # turn out to be useful (Meta sometimes sends body-only test pings).
    _record_webhook_sync()

    try:
        body = request.get_json(force=True, silent=True) or {}
    except Exception:  # noqa: BLE001
        body = {}

    # Compact body preview for the log — Meta payloads are small.
    try:
        body_preview = json.dumps(body)[:600]
    except Exception:  # noqa: BLE001
        body_preview = "<unserializable>"
    print(f"[webhook-callback] body={body_preview}", flush=True)

    touched_count    = 0  # status callbacks that updated a row
    inbound_count    = 0  # user messages that advanced an awaiting_reply row
    # Per Meta docs the payload is:
    #   entry[].changes[].value.statuses[]   ← delivery status callbacks
    #   entry[].changes[].value.messages[]   ← inbound messages from users
    # We process both halves: statuses drive the existing state machine;
    # messages are the trigger that advances a two-step row from
    # `Awaiting Reply` to `Queued` (stage=media).
    for entry in (body.get("entry") or []):
        for change in (entry.get("changes") or []):
            value = change.get("value") or {}

            # ---- Inbound user messages (two-step trigger) -------------
            for msg in (value.get("messages") or []):
                msg_from = msg.get("from") or ""
                msg_id   = msg.get("id") or ""
                msg_type = msg.get("type") or "unknown"
                # Extract a short preview for the log only.
                preview = ""
                if msg_type == "text":
                    preview = ((msg.get("text") or {}).get("body") or "")[:80]
                elif msg_type == "button":
                    preview = ((msg.get("button") or {}).get("text") or "")[:80]
                elif msg_type == "interactive":
                    inter = msg.get("interactive") or {}
                    preview = ((inter.get("button_reply") or inter.get("list_reply") or {}).get("title") or "")[:80]

                phone_digits = "".join(ch for ch in msg_from if ch.isdigit())
                print(
                    f"[inbound] from={msg_from} digits={phone_digits} type={msg_type} "
                    f"id={msg_id} preview={preview!r}",
                    flush=True,
                )
                _record_two_step_event(
                    "last_inbound",
                    phone=msg_from,
                    digits=phone_digits,
                    msg_id=msg_id,
                    msg_type=msg_type,
                    preview=preview,
                )

                # ANY inbound message counts as engagement — text, button
                # reply, interactive choice, etc. Meta opens the 24h
                # window on any inbound, so we don't filter by content.
                advanced = _advance_to_media_stage(phone_digits, msg_id)
                inbound_count += advanced

            # ---- Status callbacks (existing flow) ---------------------
            for st in (value.get("statuses") or []):
                wamid       = st.get("id")
                meta_status = (st.get("status") or "").lower()
                recipient   = st.get("recipient_id")
                try:
                    ts_s = int(st.get("timestamp")) if st.get("timestamp") else None
                except (TypeError, ValueError):
                    ts_s = None
                err_msg = None
                errs = st.get("errors") or []
                if errs:
                    err_msg = (errs[0].get("title") or errs[0].get("message")
                               or str(errs[0]))[:300]

                # Spec-required per-status log line.
                print(
                    f"[meta-status] wamid={wamid} status={meta_status} "
                    f"recipient={recipient} timestamp={ts_s} error={err_msg}",
                    flush=True,
                )

                # Update the last-callback tracker for /deliveries/webhook-status.
                with _LAST_WEBHOOK_LOCK:
                    _LAST_WEBHOOK["received_at_ms"] = _now_ms()
                    _LAST_WEBHOOK["ip"]             = ip
                    _LAST_WEBHOOK["wamid"]          = wamid
                    _LAST_WEBHOOK["meta_status"]    = meta_status
                    _LAST_WEBHOOK["error"]          = err_msg

                if not wamid:
                    continue
                # Try main wamid match first (direct + stage-2 media).
                # If no row owns this wamid as its provider_message_id,
                # try matching it against any row's prompt_wamid — that
                # tells us this status is for a stage-1 prompt, not the
                # main delivery, and we should update prompt_status only.
                if _apply_webhook_status(wamid, meta_status, ts_s, err_msg):
                    touched_count += 1
                elif _apply_prompt_status(wamid, meta_status, err_msg):
                    touched_count += 1

    print(f"[webhook-callback] processed {touched_count} status update(s), "
          f"{inbound_count} inbound trigger(s)", flush=True)
    return jsonify({
        "status":   "success",
        "updated":  touched_count,
        "inbound":  inbound_count,
    }), 200


@app.route("/deliveries/webhook-status", methods=["GET"])
def deliveries_webhook_status():
    """Lightweight diagnostics: when did the last Meta callback arrive,
    what was it, and what URL should Meta be calling. Drives the Webhook
    Status panel in the Delivery page UI."""
    pbu = (os.environ.get("PUBLIC_BASE_URL") or "").strip()
    callback_url = (pbu.rstrip("/") + "/deliveries/whatsapp-webhook") if pbu else None
    with _LAST_WEBHOOK_LOCK:
        snap = dict(_LAST_WEBHOOK)

    # ---- Rotation signal -------------------------------------------------
    # Banner logic: show "Webhook URL changed" iff the LAST URL Meta
    # actually called differs from the URL we'd ask them to call now. As
    # soon as Meta sends a verify-GET or status-POST to the new URL,
    # _record_webhook_sync() updates the sync marker and this returns
    # False on the next poll — the banner clears without a restart.
    #
    # `last_seen_public_base_url` is kept for backward-compat with older
    # frontends but now mirrors the sync URL (which is what the panel
    # was always meant to surface).
    sync_url = _read_webhook_sync()
    if not sync_url:
        # No sync recorded yet — fall back to the startup marker so a
        # fresh install with Meta already configured doesn't false-alarm.
        try:
            startup_marker = os.path.join(PROJECT_ROOT, "data", "last-public-base-url.txt")
            if os.path.isfile(startup_marker):
                with open(startup_marker, "r", encoding="utf-8") as f:
                    sync_url = f.read().strip() or None
        except Exception:  # noqa: BLE001
            pass

    tunnel_rotated = bool(sync_url and pbu and sync_url != pbu)

    return jsonify({
        "status":               "success",
        "callback_url":         callback_url,
        "public_base_url":      pbu or None,
        "verify_token_present": bool((os.environ.get("WHATSAPP_WEBHOOK_VERIFY_TOKEN") or "").strip()),
        "app_secret_present":   bool((os.environ.get("WHATSAPP_APP_SECRET") or "").strip()),
        "provider":             _PROVIDER_NAME,
        # Last-callback breadcrumb.
        "last_callback":        snap,
        # Tunnel rotation signal — frontend shows a banner if mismatched.
        "last_seen_public_base_url": sync_url,
        "synced_at_url":        sync_url,
        "tunnel_rotated":       tunnel_rotated,
    })


@app.route("/deliveries/logs", methods=["GET"])
def deliveries_logs():
    """Return the last N entries from delivery-logs.jsonl."""
    try:
        limit = max(1, min(int(request.args.get("limit", 100)), 1000))
    except ValueError:
        limit = 100
    return jsonify({
        "status": "success",
        "items":  _read_delivery_log_tail(limit),
    })


# ---------------------------------------------------------------------------
# Cross-join helper — used by both /delivery-status and /dashboard-stats so
# they prefer real delivery state where available, falling back to filesystem-
# derived "Queued" for generated outputs that have never been enqueued.
# ---------------------------------------------------------------------------

def _materialise_delivery_view():
    """Cross-join generated outputs with delivery records by stem.

    Each returned item shape (superset of the previous /delivery-status):
      {
        id, stem, name,                       <- from filesystem
        recipient_id?, recipient_phone?,      <- from delivery (if any)
        recipient_address?, recipient_name?,
        video, image,                         <- {filename, url, size}
        status,                               <- Queued | Sending | Delivered | Failed
        attempts, last_error, provider_message_id, sentAt, deliveredAt,
        createdAt                             <- file mtime
      }
    """
    fs_items = _list_generated_items()  # already sorted newest-first

    with _DELIVERIES_LOCK:
        doc = _load_deliveries()
    # Map: stem -> most-recently-updated delivery for that stem.
    latest_by_stem = {}
    for d in doc["items"]:
        s = d.get("stem")
        if not s:
            continue
        prev = latest_by_stem.get(s)
        if prev is None or d.get("updatedAt", 0) >= prev.get("updatedAt", 0):
            latest_by_stem[s] = d

    # Cache the file scans so every row's _detect_media_for_stem doesn't
    # rescan the output dirs from scratch.
    _videos = _scan_output_dir(OUTPUT_VIDEOS, _VIDEO_OUT_EXTS_SET)
    _images = _scan_output_dir(OUTPUT_IMAGES, _IMAGE_OUT_EXTS_SET)

    merged = []
    for it in fs_items:
        dlv = latest_by_stem.get(it["id"])
        # LIVE detection — wins over any stored snapshot. This is what the
        # Delivery page reads to render the Media Type pill + filename.
        det = _detect_media_for_stem(it["id"], videos=_videos, images=_images, log=False)
        base = {
            **it,
            "detected_kind":     det["kind"],
            "detected_filename": det["filename"],
            # Also surface as `media_kind` for any caller that's been
            # reading that field directly — keeps the live truth there too.
            "media_kind":        det["kind"],
        }
        if dlv:
            merged.append({
                **base,
                "status":              dlv["status"],
                "recipient_id":        dlv.get("recipient_id"),
                "recipient_name":      dlv.get("recipient_name"),
                "recipient_phone":     dlv.get("recipient_phone"),
                "recipient_address":   dlv.get("recipient_address"),
                "delivery_id":         dlv.get("id"),
                "attempts":            dlv.get("attempts", 0),
                "max_attempts":        dlv.get("max_attempts", _MAX_ATTEMPTS),
                "last_error":          dlv.get("last_error"),
                "provider":            dlv.get("provider"),
                "provider_message_id": dlv.get("provider_message_id"),
                "sentAt":              dlv.get("sentAt"),
                "deliveredAt":         dlv.get("deliveredAt"),
                # The stored kind on the delivery row — useful for debugging
                # but the UI should read `detected_kind` for display.
                "stored_media_kind":   dlv.get("media_kind"),
            })
        else:
            merged.append(base)  # filesystem-only fallback (Queued / Failed by pair)

    counts = {"Delivered": 0, "Sending": 0, "Queued": 0, "Failed": 0}
    for m in merged:
        s = m.get("status")
        if s in counts:
            counts[s] += 1
    return merged, counts


# Start the worker on app boot.
_start_worker()


# =============================================================================
# AUTH — OTP-based password reset.
#
# Flow:
#   1. POST /auth/request-otp  { email }
#        - rate-limited (3/h per IP)
#        - if email matches AUTH_EMAIL: generate 6-digit OTP, hash with salt,
#          store with 5-min expiry, send via Gmail SMTP
#        - response NEVER reveals whether the email matched (anti-enumeration)
#
#   2. POST /auth/verify-otp  { email, otp }
#        - verifies hashed OTP, consumes it on success, mints a one-time
#          reset_token valid for 5 minutes
#
#   3. POST /auth/reset-password  { reset_token, new_password }
#        - validates token + password rules
#        - on success the frontend rotates its stored credential and clears
#          the session, forcing re-login with the new password
#
# Credentials live in api/.env (gitignored). Logging goes to the Flask
# logger so SMTP success/failure is visible in the dev console.
# =============================================================================

import hashlib
import secrets
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


def _load_env_file():
    """Tiny .env loader so we don't need python-dotenv as a dependency."""
    if os.environ.get("VERCEL") == "1":
        print("[app] Vercel serverless detected, skipping late .env file load", flush=True)
        return
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.isfile(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                # Strip surrounding quotes if present.
                if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
                    v = v[1:-1]
                if k and k not in os.environ:
                    os.environ[k] = v
    except Exception as e:  # pragma: no cover
        app.logger.warning("AUTH: could not load .env — %s", e)


_load_env_file()

SMTP_HOST     = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER     = os.environ.get("SMTP_USER", "")
# Strip whitespace — Gmail App Passwords are presented with spaces.
SMTP_PASS     = os.environ.get("SMTP_PASS", "").replace(" ", "")
SMTP_FROM     = os.environ.get("SMTP_FROM", SMTP_USER) or SMTP_USER
ALLOWED_EMAIL = (os.environ.get("AUTH_EMAIL", "abacusclass963@gmail.com") or "").strip().lower()

# In-memory stores. For a single-tenant demo this is fine; production
# should move these to Redis or a small DB so multi-worker setups
# don't lose state.
_OTP_STORE       = {}            # email -> {hash, salt, expires_at, attempts_left}
_OTP_LOCK        = threading.Lock()
_RESET_TOKENS    = {}            # token -> {email, expires_at}
_RESET_LOCK      = threading.Lock()
_RATE_LIMITS     = {}            # "action:ip" -> [timestamps]
_RATE_LOCK       = threading.Lock()


def _client_ip():
    raw = request.headers.get("X-Forwarded-For") or request.remote_addr or "unknown"
    return raw.split(",")[0].strip()


def _check_rate(action, max_per_hour=3, window_sec=3600):
    """LocalIP-keyed rate limit. Returns (allowed: bool, retry_after_min)."""
    ip  = _client_ip()
    key = f"{action}:{ip}"
    with _RATE_LOCK:
        now = time.time()
        bucket = [t for t in _RATE_LIMITS.get(key, []) if now - t < window_sec]
        if len(bucket) >= max_per_hour:
            oldest = min(bucket)
            retry_after_min = max(1, int((window_sec - (now - oldest)) / 60) + 1)
            _RATE_LIMITS[key] = bucket
            return False, retry_after_min
        bucket.append(now)
        _RATE_LIMITS[key] = bucket
        return True, None


def _hash_otp(otp, salt):
    return hashlib.sha256((salt + ":" + otp).encode("utf-8")).hexdigest()


def _password_meets_rules(pw):
    if not pw or len(pw) < 8:
        return False
    if not re.search(r"[A-Z]", pw):  return False
    if not re.search(r"[a-z]", pw):  return False
    if not re.search(r"\d", pw):     return False
    if not re.search(r"[^a-zA-Z0-9]", pw): return False
    return True


def _otp_email_html(otp):
    return f"""\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Verification Code</title>
</head>
<body style="margin:0;padding:0;background:#FFF7EC;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FFF7EC;padding:40px 16px;">
    <tr><td align="center">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="max-width:540px;background:#FFFFFF;border-radius:18px;overflow:hidden;box-shadow:0 24px 60px rgba(194,65,12,0.10);">
        <tr><td style="background:linear-gradient(135deg,#F97316 0%,#EA580C 100%);padding:32px 32px 24px;">
          <p style="margin:0;font-size:11px;font-weight:700;letter-spacing:0.16em;text-transform:uppercase;color:rgba(255,255,255,0.78);">Mastermind Automation</p>
          <h1 style="margin:6px 0 0;font-size:22px;font-weight:800;letter-spacing:-0.02em;color:#FFFFFF;">Verification Code</h1>
        </td></tr>
        <tr><td style="padding:36px 32px 8px;">
          <p style="margin:0 0 24px;font-size:15px;line-height:1.6;color:#4A3A2E;">
            Hi there,<br><br>
            You requested a verification code to reset your Mastermind Automation Studio password. Use the code below to continue.
          </p>
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
            <tr><td align="center" style="padding:28px;background:linear-gradient(135deg,#FFF7E7 0%,#FFE9CC 100%);border:1px solid rgba(180,83,9,0.14);border-radius:14px;">
              <p style="margin:0;font-size:11px;font-weight:700;letter-spacing:0.18em;text-transform:uppercase;color:#C2410C;">Your Code</p>
              <p style="margin:12px 0 0;font-size:42px;font-weight:800;letter-spacing:0.36em;color:#1A1410;font-family:'SF Mono',Menlo,Consolas,monospace;">{otp}</p>
            </td></tr>
          </table>
          <p style="margin:24px 0 0;font-size:13.5px;line-height:1.6;color:#6B5C50;">
            This code <strong style="color:#1A1410;">expires in 5 minutes</strong> and can be used once.
          </p>
          <p style="margin:16px 0 0;font-size:13px;line-height:1.6;color:#8C7164;">
            If you did not request this code, you can safely ignore this email — your password will remain unchanged.
          </p>
        </td></tr>
        <tr><td style="padding:24px 32px 32px;border-top:1px solid rgba(180,83,9,0.08);">
          <p style="margin:0;font-size:12px;color:#8C7164;font-weight:500;">
            Mastermind Abacus &middot; Automation Studio<br>
            <span style="color:#B5A290;">Built for modern automation teams.</span>
          </p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""


def _send_otp_email(to_email, otp):
    """Send the OTP via Gmail SMTP. Returns True on success, False otherwise.
    Logs SMTP status to the Flask logger."""
    if not SMTP_USER or not SMTP_PASS:
        app.logger.error("SMTP: credentials missing — set SMTP_USER and SMTP_PASS in api/.env")
        return False

    text = (
        f"Your Mastermind Automation verification code is: {otp}\n\n"
        "This code expires in 5 minutes. If you did not request it, ignore this email."
    )
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "Mastermind Automation • Verification Code"
    msg["From"]    = SMTP_FROM
    msg["To"]      = to_email
    msg.attach(MIMEText(text, "plain", "utf-8"))
    msg.attach(MIMEText(_otp_email_html(otp), "html", "utf-8"))

    try:
        app.logger.info("SMTP: connecting to %s:%d as %s", SMTP_HOST, SMTP_PORT, SMTP_USER)
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls(context=ssl.create_default_context())
            smtp.ehlo()
            smtp.login(SMTP_USER, SMTP_PASS)
            smtp.sendmail(SMTP_FROM, [to_email], msg.as_string())
        app.logger.info("SMTP: OTP email sent to %s", to_email)
        return True
    except smtplib.SMTPAuthenticationError as e:
        app.logger.error(
            "SMTP: authentication failed (%s). Gmail requires a 16-char App Password, "
            "not your account password. Generate one at https://myaccount.google.com/apppasswords",
            e,
        )
        return False
    except (smtplib.SMTPException, OSError) as e:
        app.logger.error("SMTP: send failed — %s", e)
        return False


# ----- routes ----------------------------------------------------------------
@app.route("/auth/request-otp", methods=["POST"])
def auth_request_otp():
    data  = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()

    if not email or "@" not in email:
        return jsonify({"ok": False, "error": "invalid_email"}), 400

    allowed, retry_after = _check_rate("forgot", max_per_hour=3)
    if not allowed:
        app.logger.info("AUTH: rate-limit hit for %s from %s", email, _client_ip())
        return jsonify({"ok": False, "error": "rate_limit",
                        "retry_after_min": retry_after}), 429

    # Only actually generate and send when the email matches the registered
    # account. Otherwise return the same generic response (anti-enumeration).
    if email == ALLOWED_EMAIL:
        otp  = f"{secrets.randbelow(1000000):06d}"
        salt = secrets.token_hex(16)
        with _OTP_LOCK:
            _OTP_STORE[email] = {
                "hash":          _hash_otp(otp, salt),
                "salt":          salt,
                "expires_at":    time.time() + 5 * 60,
                "attempts_left": 3,
            }
        app.logger.info("AUTH: OTP generated for %s (expires in 5 min)", email)
        sent = _send_otp_email(email, otp)
        if not sent:
            # Drop the OTP — user didn't actually receive it.
            with _OTP_LOCK:
                _OTP_STORE.pop(email, None)
            return jsonify({"ok": False, "error": "smtp_failed"}), 502
    else:
        app.logger.info("AUTH: OTP request for non-registered email (anti-enumeration)")

    return jsonify({
        "ok": True,
        "message": "If the email exists, a verification code has been sent.",
    })


@app.route("/auth/verify-otp", methods=["POST"])
def auth_verify_otp():
    data  = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    otp   = (data.get("otp")   or "").strip()

    if not otp.isdigit() or len(otp) != 6:
        return jsonify({"ok": False, "error": "invalid"}), 400

    with _OTP_LOCK:
        entry = _OTP_STORE.get(email)
        if not entry:
            return jsonify({"ok": False, "error": "expired"}), 400
        if time.time() > entry["expires_at"]:
            _OTP_STORE.pop(email, None)
            return jsonify({"ok": False, "error": "expired"}), 400
        if entry["attempts_left"] <= 0:
            _OTP_STORE.pop(email, None)
            return jsonify({"ok": False, "error": "attempts"}), 400
        if _hash_otp(otp, entry["salt"]) != entry["hash"]:
            entry["attempts_left"] -= 1
            app.logger.info("AUTH: OTP mismatch for %s — %d attempts left",
                            email, entry["attempts_left"])
            return jsonify({"ok": False, "error": "invalid",
                            "attempts_left": entry["attempts_left"]}), 400
        # Success — consume the OTP and mint a one-time reset token.
        _OTP_STORE.pop(email, None)
        token = secrets.token_urlsafe(32)
        with _RESET_LOCK:
            _RESET_TOKENS[token] = {
                "email":      email,
                "expires_at": time.time() + 5 * 60,
            }

    app.logger.info("AUTH: OTP verified for %s; reset token issued", email)
    return jsonify({"ok": True, "reset_token": token})


@app.route("/auth/reset-password", methods=["POST"])
def auth_reset_password():
    data         = request.get_json(silent=True) or {}
    token        = (data.get("reset_token")  or "").strip()
    new_password =  data.get("new_password") or ""

    if not _password_meets_rules(new_password):
        return jsonify({"ok": False, "error": "weak"}), 400

    with _RESET_LOCK:
        entry = _RESET_TOKENS.pop(token, None)
    if not entry:
        return jsonify({"ok": False, "error": "invalid_token"}), 400
    if time.time() > entry["expires_at"]:
        return jsonify({"ok": False, "error": "invalid_token"}), 400

    # In a real backend the new password would be bcrypt-hashed and stored in
    # a users table here. For this single-tenant demo the frontend rotates
    # its localStorage credential after we confirm the token is valid.
    app.logger.info("AUTH: password reset completed for %s", entry["email"])
    return jsonify({"ok": True})


if __name__ == "__main__":
    print("[api-ready] flask app.py listening on http://127.0.0.1:5000", flush=True)
    app.run(host="127.0.0.1", port=5000, debug=False)
